import openpyxl
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.line import LineProperties
from collections import Counter
from consts import mbti_colors, MBTI_TYPES, MBTI_LETTERS


def create_distribution_charts(workbook):
    # Check if 'MBTI Distribution' sheet already exists, if not create it
    if 'MBTI Distribution' in workbook.sheetnames:
        chart_sheet = workbook['MBTI Distribution']
    else:
        chart_sheet = workbook.create_sheet(title="MBTI Distribution")
    
    # Create the main distribution chart
    create_main_distribution_chart(chart_sheet)
    
    # Create dichotomy charts
    create_dichotomy_charts(chart_sheet)

    # Analyze internal and external components and create pie charts
    analyze_mbti_external_internal(chart_sheet)

    # Adjust column widths
    adjust_column_widths(chart_sheet)


def create_main_distribution_chart(chart_sheet):
    # Write headers
    chart_sheet['A40'] = "MBTI Type Data"
    chart_sheet['A40'].font = Font(bold=True)

    # Write MBTI types and count formulas
    for row, mbti_type in enumerate(MBTI_TYPES, start=41):
        chart_sheet[f'A{row}'] = mbti_type
        chart_sheet[f'B{row}'] = f"=COUNTIF(Table1[Type],A{row})"
        if mbti_type in mbti_colors:
            chart_sheet[f'A{row}'].fill = PatternFill(start_color=mbti_colors[mbti_type],
                                                     end_color=mbti_colors[mbti_type],
                                                     fill_type="solid")

    # Create pie chart
    main_chart = PieChart()
    main_chart.title = "Distribution by Type"

    labels = Reference(chart_sheet, min_col=1, min_row=41, max_row=56)
    data = Reference(chart_sheet, min_col=2, min_row=41, max_row=56)
    main_chart.add_data(data, titles_from_data=False)
    main_chart.set_categories(labels)
    series = main_chart.series[0]
    series.data_points = []
    # Set colors for each slice
    for i, mbti_type in enumerate(mbti_colors.keys()):
        color = mbti_colors[mbti_type]
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = color
        series.data_points.append(dp)
        print(f"MBTI Type: {mbti_type}, Color: {color}")
    # Chart size
    main_chart.width = 20  # wider for readability
    main_chart.height = 20

    # Data labels: Only show MBTI type, not value or percent
    main_chart.dataLabels = DataLabelList()
    main_chart.dataLabels.showCatName = True
    main_chart.dataLabels.showVal = False
    main_chart.dataLabels.showPercent = True
    main_chart.dataLabels.showSerName = False

    # Add the chart to the sheet
    chart_sheet.add_chart(main_chart, "A1")


def create_dichotomy_charts(chart_sheet):
    dichotomies = [
        ("E", "I", "Energy source - E/I"),
        ("S", "N", "Information - S/N"),
        ("T", "F", "Decisions - T/F"),
        ("J", "P", "Lifestyle - J/P")
    ]
    chart_sheet['D40'] = "Dichotomies Data"
    chart_sheet['D40'].font = Font(bold=True)
    chart_sheet['D41'] = "Energy source - E/I"
    chart_sheet['D42'] = "Extroversion"
    chart_sheet['D43'] = "Introversion"
    chart_sheet['D44'] = "Information - S/N"
    chart_sheet['D45'] = "Sensing"
    chart_sheet['D46'] = "Intuition"
    chart_sheet['D47'] = "Decisions - T/F"
    chart_sheet['D48'] = "Thinking"
    chart_sheet['D49'] = "Feeling"
    chart_sheet['D50'] = "Lifestyle - J/P"
    chart_sheet['D51'] = "Judging"
    chart_sheet['D52'] = "Perceiving"
    chart_sheet['E42'] = "=COUNTIF(Table1[Type], \"*E*\")-1"
    chart_sheet['E43'] = "=COUNTIF(Table1[Type], \"*I*\")"
    chart_sheet['E45'] = "=COUNTIF(Table1[Type], \"*S*\")"
    chart_sheet['E46'] = "=COUNTIF(Table1[Type], \"*N*\")"
    chart_sheet['E48'] = "=COUNTIF(Table1[Type], \"*T*\")"
    chart_sheet['E49'] = "=COUNTIF(Table1[Type], \"*F*\")"
    chart_sheet['E51'] = "=COUNTIF(Table1[Type], \"*J*\")"
    chart_sheet['E52'] = "=COUNTIF(Table1[Type], \"*P*\")"
    # general chart data for all dichotomies
    pie_width = 12
    pie_height = 6.5

    # Create E/I pie chart
    pie_EI = create_dichotomy_pie_chart(chart_sheet, "Energy source - E/I", 42, 43, pie_width, pie_height)
    chart_sheet.add_chart(pie_EI, "H1")

    # Create S/N pie chart
    pie_SN = create_dichotomy_pie_chart(chart_sheet, "Information - S/N", 45, 46, pie_width, pie_height)
    chart_sheet.add_chart(pie_SN, "H14")

    # Create T/F pie chart
    pie_TF = create_dichotomy_pie_chart(chart_sheet, "Decisions - T/F", 48, 49, pie_width, pie_height)
    chart_sheet.add_chart(pie_TF, "L1")

    # Create J/P pie chart
    pie_JP = create_dichotomy_pie_chart(chart_sheet, "Lifestyle - J/P", 51, 52, pie_width, pie_height)
    chart_sheet.add_chart(pie_JP, "L14")


def create_dichotomy_pie_chart(chart_sheet, title, start_row, end_row, width, height):
    pie = PieChart()
    pie.title = title
    labels = Reference(chart_sheet, min_col=4, min_row=start_row, max_row=end_row)
    data = Reference(chart_sheet, min_col=5, min_row=start_row, max_row=end_row)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.width = width
    pie.height = height

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showPercent = True
    pie.dataLabels.showSerName = False

    return pie

def analyze_mbti_external_internal(chart_sheet):
    External_list = ["IJ", "IP", "EJ", "EP", ]
    Internal_list = ["ST", "SF", "NF", "NT", ]
    start_row = 41
    end_row = 56
    chart_sheet['G40'] = "Internal Analysis"
    chart_sheet['J40'] = "External Analysis"
    chart_sheet['G40'].font = Font(bold=True)
    chart_sheet['J40'].font = Font(bold=True)
    chart_sheet['G41'] = "Internal"
    chart_sheet['J41'] = "External"
    chart_sheet['H41'] = "Count"
    chart_sheet['K41'] = "Count"
    chart_sheet['H42'] = '=COUNTIF(Table1[Type], "*ST*")'
    chart_sheet['H43'] = '=COUNTIF(Table1[Type], "*SF*")'
    chart_sheet['H44'] = '=COUNTIF(Table1[Type], "*NF*")'
    chart_sheet['H45'] = '=COUNTIF(Table1[Type], "*NT*")'
    chart_sheet['K42'] = '=COUNTIF(Table1[Type], "I*J")'
    chart_sheet['K43'] = '=COUNTIF(Table1[Type], "I*P")'
    chart_sheet['K44'] = '=COUNTIF(Table1[Type], "E*J")'
    chart_sheet['K45'] = '=COUNTIF(Table1[Type], "E*P")'

    # Insert Internal_list at G42->G45
    for i, internal_type in enumerate(Internal_list):
        chart_sheet[f'G{42+i}'] = internal_type

    # Insert External_list at J42->J45
    for i, external_type in enumerate(External_list):
        chart_sheet[f'J{42+i}'] = external_type

    # Create Internal Analysis Pie Chart
    internal_pie = PieChart()
    internal_pie.title = "Internal Analysis Distribution"
    labels = Reference(chart_sheet, min_col=7, min_row=42, max_row=45)
    data = Reference(chart_sheet, min_col=8, min_row=42, max_row=45)
    internal_pie.add_data(data, titles_from_data=False)
    internal_pie.set_categories(labels)
    internal_pie.width = 12
    internal_pie.height = 6.5

    # Create External Analysis Pie Chart
    external_pie = PieChart()
    external_pie.title = "External Analysis Distribution"
    labels = Reference(chart_sheet, min_col=10, min_row=42, max_row=45)
    data = Reference(chart_sheet, min_col=11, min_row=42, max_row=45)
    external_pie.add_data(data, titles_from_data=False)
    external_pie.set_categories(labels)
    external_pie.width = 12
    external_pie.height = 6.5

    # Configure data labels for both charts
    for pie in [internal_pie, external_pie]:
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showPercent = True
        pie.dataLabels.showSerName = False

    # Add the charts to the sheet
    chart_sheet.add_chart(internal_pie, "H27")
    chart_sheet.add_chart(external_pie, "L27")

def adjust_column_widths(chart_sheet):
    for column in chart_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        chart_sheet.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    # For testing purposes
    import openpyxl
    workbook_path = r"F:\projects\MBTInfo\output\MBTI_Results.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    create_distribution_charts(workbook)
    workbook.save(workbook_path)
