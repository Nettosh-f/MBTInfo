import openpyxl as xl
from openpyxl.styles import PatternFill, Color, Font
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from consts import mbti_colors


def adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width


def apply_formatting(workbook_path):
    workbook = xl.load_workbook(workbook_path)
    sheet = workbook.active

    # Adjust column widths
    adjust_column_widths(sheet)

    # Define colors
    soft_red = Color(rgb="FFC7CE")  # Light red
    soft_yellow = Color(rgb="FFEB9C")  # Light yellow
    soft_green = Color(rgb="C6EFCE")  # Light green

    # Color scale rule for MBTI scores (columns D to K)
    scores_range = f"D2:K{sheet.max_row}"

    # Rule for black color when value is 0
    zero_rule = Rule(
        type="cellIs",
        operator="equal",
        formula=["0"],
        stopIfTrue=True,
        dxf=DifferentialStyle(
            fill=PatternFill(start_color="000000", end_color="000000", fill_type="solid"),
            font=Font(color="FFFFFF")  # White text for better visibility
        )
    )
    sheet.conditional_formatting.add(scores_range, zero_rule)

    # Color scale rule for non-zero values
    color_scale = ColorScaleRule(
        start_type='num', start_value=1, start_color="FFC7CE",  # Soft red for low values
        mid_type='num', mid_value=15, mid_color="FFEB9C",  # Soft yellow for mid values
        end_type='num', end_value=30, end_color="C6EFCE"  # Soft green for high values
    )
    sheet.conditional_formatting.add(scores_range, color_scale)

    # Formatting rules for facets (columns L onwards)
    in_fill = PatternFill(start_color="C0CEE6", end_color="C0CEE6", fill_type="solid")
    mid_fill = PatternFill(start_color="D6E1C5", end_color="D6E1C5", fill_type="solid")
    out_fill = PatternFill(start_color="B9CA9D", end_color="B9CA9D", fill_type="solid")
    dash_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    facets_range = f"L2:{xl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"
    
    in_rule = Rule(type="cellIs", operator="equal", formula=['"IN-PREF"'],
                   stopIfTrue=False, dxf=DifferentialStyle(fill=in_fill))
    mid_rule = Rule(type="cellIs", operator="equal", formula=['"MIDZONE"'],
                    stopIfTrue=False, dxf=DifferentialStyle(fill=mid_fill))
    out_rule = Rule(type="cellIs", operator="equal", formula=['"OUT-OF-PREF"'],
                    stopIfTrue=False, dxf=DifferentialStyle(fill=out_fill))
    dash_rule = Rule(type="cellIs", operator="equal", formula=['-'],
                     stopIfTrue=False, dxf=DifferentialStyle(fill=dash_fill))
    sheet.conditional_formatting.add(facets_range, in_rule)
    sheet.conditional_formatting.add(facets_range, mid_rule)
    sheet.conditional_formatting.add(facets_range, out_rule)

    # Apply conditional formatting for MBTI types in column C
    mbti_range = f"C2:C{sheet.max_row}"
    for mbti_type, color in mbti_colors.items():
        mbti_rule = Rule(
            type="cellIs",
            operator="equal",
            formula=[f'"{mbti_type}"'],
            stopIfTrue=False,
            dxf=DifferentialStyle(fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
        )
        sheet.conditional_formatting.add(mbti_range, mbti_rule)

    workbook.save(workbook_path)
    print(f"Formatting applied to {workbook_path}")


if __name__ == "__main__":
    # For testing purposes
    apply_formatting(r"F:\projects\MBTInfo\output\MBTI_Results.xlsx")