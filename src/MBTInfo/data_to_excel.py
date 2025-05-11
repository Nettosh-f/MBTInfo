import os
import openpyxl as xl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font
from utils import (get_all_info, find_and_parse_mbti_scores, convert_scores_to_mbti_dict, collect_qualities,
                   check_communication, check_managing_change, check_managing_conflict)
from consts import FACETS

def process_pdf_to_xl(text_path, output_dir, result_sheet_name, output_filename):
    info = get_all_info(text_path)
    qualities = find_and_parse_mbti_scores(text_path)
    mbti_dict = convert_scores_to_mbti_dict(qualities)
    preferred_qualities, midzone_qualities, out_qualities = collect_qualities(text_path)

    # Get communication, change management, and conflict management facets
    communication_facets = check_communication(text_path)
    change_facets = check_managing_change(text_path)
    conflict_facets = check_managing_conflict(text_path)

    # Convert all qualities to lowercase for case-insensitive comparison
    preferred_qualities = [q.lower() for q in preferred_qualities]
    midzone_qualities = [q.lower() for q in midzone_qualities]
    out_qualities = [q.lower() for q in out_qualities]

    output_path = os.path.join(output_dir, output_filename)

    headers = ["Name", "Date", "Type", "Extroversion", "Introversion", "Sensing", "Intuition", "Thinking", "Feeling", "Judging", "Perceiving"]
    headers.extend(FACETS)  # Add all facets to the headers
    
    sections = ["Communicating", "Managing Change", "Managing Conflict"]
    headers.extend(sections)  # Add section headers without empty cells
    
    # Calculate the actual number of columns needed
    actual_columns = len(headers)
    # Add extra columns for the section data (8 empty cells after each section header)
    actual_columns += 8 * len(sections)
    
    if os.path.exists(output_path):
        workbook = xl.load_workbook(filename=str(output_path))
        if result_sheet_name in workbook.sheetnames:
            sheet = workbook[result_sheet_name]
            
            # Find the last non-empty row
            last_row = sheet.max_row
            while last_row > 1 and all(cell.value is None for cell in sheet[last_row]):
                last_row -= 1
            last_row += 1  # Move to the next empty row
        else:
            sheet = workbook.create_sheet(result_sheet_name)
            _setup_headers(sheet, headers, sections)
            last_row = 2  # Start data from the second row
    else:
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = result_sheet_name
        _setup_headers(sheet, headers, sections)
        last_row = 2  # Start data from the second row

    data = [info['name'], info['date'], info['type']] + list(mbti_dict.values())
    
    # Add values for facets
    for header in headers[11:-3]:  # Exclude the last 3 sections
        header_lower = header.lower()
        if header_lower in preferred_qualities:
            data.append('IN-PREF')
        elif header_lower in midzone_qualities:
            data.append('MIDZONE')
        elif header_lower in out_qualities:
            data.append('OUT-OF-PREF')
        else:
            data.append('-')
    
    # Add values for communication, change management, and conflict management
    data.extend(communication_facets + [''] * (9 - len(communication_facets)))
    data.extend(change_facets + [''] * (9 - len(change_facets)))
    data.extend(conflict_facets + [''] * (9 - len(conflict_facets)))
    
    # Append data to the last empty row
    for col, value in enumerate(data, start=1):
        sheet.cell(row=last_row, column=col, value=value if value else '')

    # Update or create the table
    last_column = xl.utils.get_column_letter(actual_columns)
    table_range = f"A1:{last_column}{sheet.max_row}"
    
    # If table exists, update its reference; otherwise create a new one
    if 'Table1' in sheet.tables:
        # Store the table style before removing
        old_table = sheet.tables['Table1']
        style_info = old_table.tableStyleInfo
        
        # Remove the old table
        del sheet.tables['Table1']
        
        # Create a new table with the updated range but same style
        tab = Table(displayName="Table1", ref=table_range)
        tab.tableStyleInfo = style_info
        sheet.add_table(tab)
    else:
        # Create a new table
        tab = Table(displayName="Table1", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)
    
    workbook.save(filename=str(output_path))
    return output_path


def _unmerge_first_row(sheet):
    for merge_range in list(sheet.merged_cells.ranges):
        if merge_range.min_row == 1:
            sheet.unmerge_cells(str(merge_range))


def _setup_headers(sheet, headers, sections):
    # Unmerge any existing merged cells in the first row
    for merge_range in list(sheet.merged_cells.ranges):
        if merge_range.min_row == 1:
            sheet.unmerge_cells(str(merge_range))

    # Set up the headers in a single row
    col = 1
    for header in headers:
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        
        if header in sections:
            cell.alignment = Alignment(horizontal='center')
            # Add 8 empty cells after the section header
            for _ in range(8):
                col += 1
                sheet.cell(row=1, column=col).value = ''
        col += 1

    # Adjust column widths to fit content
    for column_cells in sheet.columns:
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value]
        if lengths:
            max_length = max(lengths)
            sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2
        else:
            # Set a default width for empty columns
            sheet.column_dimensions[column_cells[0].column_letter].width = 10

    # Freeze the first row
    sheet.freeze_panes = 'A2'

if __name__ == "__main__":
    pdf_path = r"F:\projects\MBTInfo\output\textfiles\nir-bensinai-MBTI_text.txt"
    output_dir = r"F:\projects\MBTInfo\output"
    output_filename = "MBTI_Results_test.xlsx"
    result_sheet_name = "MBTI Results"
    output_path = process_pdf_to_xl(pdf_path, output_dir, result_sheet_name, output_filename)
    print(f"MBTI results appended to {output_path}")
