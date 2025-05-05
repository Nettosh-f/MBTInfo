import openpyxl as xl
from openpyxl.styles import PatternFill, Color, Font
from openpyxl.formatting.rule import ColorScaleRule, Rule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from consts import mbti_colors, FACETS, MIDZONE_FACETS


def adjust_column_widths(sheet):
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, xl.cell.cell.MergedCell):
                continue
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width


def format_xl(file_path):
    workbook = xl.load_workbook(file_path)
    sheet = workbook.active
    adjust_column_widths(sheet)
    format_headers(sheet)

    # colors:
    black_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    light_green = 'ffb7e5b7'
    light_yellow = 'ffe5e589'
    light_red = 'ffe5b7b7'

    # handling MBTI type colors
    for sheet in workbook.worksheets:
        print(f"Applying MBTI rules to sheet: {sheet.title}")
        for mbti_type, color in mbti_colors.items():
            mbti_rule = Rule(
                type="cellIs",
                operator="equal",
                formula=[f'"{mbti_type}"'],
                stopIfTrue=False,
                dxf=DifferentialStyle(fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
            )
            # Apply the rule to the entire sheet
            mbti_range = sheet.dimensions
            sheet.conditional_formatting.add(mbti_range, mbti_rule)
        print(f"MBTI rules applied to sheet: {sheet.title}")

    # The rest of the formatting applies only to the active sheet
    sheet = workbook.active
    print(f"Applying additional formatting to active sheet: {sheet.title}")

    # handling qualities scores
    scores_range = f"D2:K{sheet.max_row}"
    formula = 'OR(D2<1,D2>30)'
    zero_rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=black_fill)
    sheet.conditional_formatting.add(scores_range, zero_rule)
    color_scale_rule = ColorScaleRule(
        start_type='num', start_value=1, start_color=light_green,
        mid_type='num', mid_value=15, mid_color=light_yellow,
        end_type='num', end_value=30, end_color=light_red,
    )
    sheet.conditional_formatting.add(scores_range, color_scale_rule)

    # handling in-preference formatting
    inpref_fill = PatternFill(start_color='FFC0CEE6', end_color='FFC0CEE6', fill_type='solid')
    midzone_fill = PatternFill(start_color='FFD6E1C5', end_color='FFD6E1C5', fill_type='solid')
    outofpref_fill = PatternFill(start_color='FFCC66', end_color='FFCC66', fill_type='solid')
    dash_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    facets_range = f"L2:AY{sheet.max_row}"
    inpref_rule = FormulaRule(formula=['L2="IN-PREF"'], fill=inpref_fill)
    midzone_rule = FormulaRule(formula=['L2="MIDZONE"'], fill=midzone_fill)
    outofpref_rule = FormulaRule(formula=['L2="OUT-OF-PREF"'], fill=outofpref_fill)
    dash_rule = FormulaRule(formula=['L2="-"'], fill=dash_fill)
    sheet.conditional_formatting.add(facets_range, inpref_rule)
    sheet.conditional_formatting.add(facets_range, midzone_rule)
    sheet.conditional_formatting.add(facets_range, outofpref_rule)
    sheet.conditional_formatting.add(facets_range, dash_rule)
    facet_format(sheet, workbook)
    workbook.save(file_path)
    print(f"Formatting applied to {file_path} successfully.")


def facet_format(main_sheet, workbook):
    max_row = main_sheet.max_row
    fill_colors = {
        'IN-PREF': 'FFC0CEE6',
        'MIDZONE': 'FFD6E1C5',
        'OUT-OF-PREF': 'FFCC66'
    }

    # Create a dictionary to map individual facets to their midzone pairs
    midzone_pairs = {}
    for pair in MIDZONE_FACETS:
        facet1, facet2 = pair.split('–')
        midzone_pairs[facet1.lower()] = pair
        midzone_pairs[facet2.lower()] = pair

    rules = {
        pref: {
            facet.lower(): Rule(
                type="cellIs",
                operator="equal",
                formula=[f'"{facet.lower()}"'],
                stopIfTrue=False,
                dxf=DifferentialStyle(fill=PatternFill(start_color=color, end_color=color, fill_type='solid'))
            ) for facet in FACETS
        } for pref, color in fill_colors.items()
    }

    # Add rules for midzone pairs
    for pair in MIDZONE_FACETS:
        rules['MIDZONE'][pair.lower()] = Rule(
            type="cellIs",
            operator="equal",
            formula=[f'"{pair.lower()}"'],
            stopIfTrue=False,
            dxf=DifferentialStyle(fill=PatternFill(start_color=fill_colors['MIDZONE'], end_color=fill_colors['MIDZONE'], fill_type='solid'))
        )

    for row in range(2, max_row + 1):
        facet_prefs = {}
        for col in range(12, 52):
            pref = main_sheet.cell(row=row, column=col).value
            if pref in fill_colors:
                facet_name = main_sheet.cell(row=1, column=col).value.lower()
                if pref == 'MIDZONE' and facet_name in midzone_pairs:
                    facet_prefs[midzone_pairs[facet_name].lower()] = pref
                else:
                    facet_prefs[facet_name] = pref

        fill_range = f"AZ{row}:BZ{row}"
        for facet, pref in facet_prefs.items():
            for sheet in workbook.worksheets:
                sheet.conditional_formatting.add(fill_range, rules[pref][facet])
        # Apply facet rules to the Facet Table sheet
        facet_table_sheet = workbook['Facet Table']
        facet_table_range = f"D{row}:T{row}"
        for facet, pref in facet_prefs.items():
            facet_table_sheet.conditional_formatting.add(facet_table_range, rules[pref][facet])
        section_list = ['Communicating', 'Managing Change', 'Managing Conflict']
        for item in section_list:
            table_sheet = workbook[f"{item}"]
            table_range = f"D{row + 2}:T{row + 2}"
            for facet, pref in facet_prefs.items():
                table_sheet.conditional_formatting.add(table_range, rules[pref][facet])

def format_headers(sheet):
    sheet['C1'].fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
    for col in range(4, 12):  # Columns D to K
        sheet.cell(row=1, column=col).fill = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
    for col in range(22, 32):  # Columns V to AE
        sheet.cell(row=1, column=col).fill = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
    for col in range(44, 52):  # Columns AR to AY
        sheet.cell(row=1, column=col).fill = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
    for col in range(52, 61):  # Columns AZ to BH
        sheet.cell(row=1, column=col).fill = PatternFill(start_color='66CC33', end_color='66CC33', fill_type='solid')
    for col in range(61, 70):  # Columns BI to BQ
        sheet.cell(row=1, column=col).fill = PatternFill(start_color='FFCC66', end_color='FFCC66', fill_type='solid')
    for col in range(70, 79):  # Columns BR to BZ
        sheet.cell(row=1, column=col).fill = PatternFill(start_color='999999', end_color='999999', fill_type='solid')


if __name__ == "__main__":
    # For testing purposes
    format_xl(r"F:\projects\MBTInfo\output\MBTI_Results.xlsx")