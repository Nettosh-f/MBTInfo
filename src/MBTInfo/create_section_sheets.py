import os
import openpyxl as xl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from utils import get_all_info, check_communication, check_managing_change, check_managing_conflict


def create_section_sheets(input_directory, workbook):
    sections = ["Communicating", "Managing Change", "Managing Conflict"]
    check_functions = [check_communication, check_managing_change, check_managing_conflict]

    for section, check_function in zip(sections, check_functions):
        sheet = workbook.create_sheet(title=section)
        _setup_sheet(sheet, section)

        row = 4  # Start data from row 4
        for file in os.listdir(input_directory):
            if file.endswith('_text.txt'):
                text_path = os.path.join(input_directory, file)
                info = get_all_info(text_path)
                facets = check_function(text_path)
                _add_data_row(sheet, row, info, facets)
                row += 1

        _create_table(sheet, row - 1)  # Create table after all data is added

    return workbook


def _setup_sheet(sheet, title):
    # Add title
    sheet['A1'] = title
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Add headers
    headers = ["Name", "Date", "Type"] + [f"Facet {i}" for i in range(1, 10)]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)

def _add_data_row(sheet, row, info, facets):
    data = [info['name'], info['date'], info['type']] + facets + [''] * (9 - len(facets))
    for col, value in enumerate(data, start=1):
        sheet.cell(row=row, column=col, value=value)

def _create_table(sheet, last_row):
    table_ref = f"A3:L{last_row}"  # Updated to include all 12 columns
    tab = Table(displayName=f"{sheet.title.replace(' ', '')}Table", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    # Adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2