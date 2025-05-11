import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from data_extractor import extract_and_save_text
from utils import get_all_info, find_and_parse_mbti_scores, convert_scores_to_mbti_dict, collect_qualities
from consts import FACETS, type_descriptions, career_insights


def generate_personal_report(input_pdf_path, output_dir, output_filename):
    """
    Generate a personal MBTI report from a PDF file.
    
    Args:
        input_pdf_path (str): Path to the input PDF file
        output_dir (str): Directory to save the output Excel file
        output_filename (str): Name of the output Excel file
    
    Returns:
        str: Path to the generated Excel file
    """
    # Extract text from PDF
    text_file_path = extract_and_save_text(input_pdf_path, output_dir)
    print(f"Text extracted to: {text_file_path}")
    if not text_file_path:
        raise ValueError("Failed to extract text from PDF file")
    
    # Create textfiles directory if it doesn't exist
    textfiles_dir = os.path.join(output_dir, 'textfiles')
    os.makedirs(textfiles_dir, exist_ok=True)
    
    # Read the extracted text content
    try:
        with open(text_file_path, 'r', encoding='utf-8') as f:
            text_content = f.read()
        print(f"Successfully read {len(text_content)} characters from {text_file_path}")
    except Exception as e:
        print(f"Error reading text file: {str(e)}")
        raise ValueError(f"Failed to read extracted text: {str(e)}")
    
    # Save a copy to the textfiles directory
    base_name = os.path.splitext(os.path.basename(input_pdf_path))[0]
    textfiles_copy_path = os.path.join(textfiles_dir, f"{base_name}_text.txt")
    
    if text_file_path != textfiles_copy_path:  # Only copy if paths are different
        try:
            with open(textfiles_copy_path, 'w', encoding='utf-8') as f:
                f.write(text_content)
            print(f"Text content copied to: {textfiles_copy_path}")
        except Exception as e:
            print(f"Warning: Failed to copy text to textfiles directory: {str(e)}")
            # Continue execution even if copy fails
    
    # Extract MBTI information
    info = get_all_info(text_file_path)
    print(f"MBTI Information: {info}")
    if not info:
        info = {'name': 'Unknown', 'date': 'Unknown', 'type': 'Unknown'}
    
    qualities = find_and_parse_mbti_scores(text_file_path)
    if not qualities:
        qualities = {}
    
    mbti_dict = convert_scores_to_mbti_dict(qualities)
    if not mbti_dict:
        # Create a default dictionary with zeros
        mbti_dict = {'E': 0, 'I': 0, 'S': 0, 'N': 0, 'T': 0, 'F': 0, 'J': 0, 'P': 0}
    
    # Handle potential None values from collect_qualities
    try:
        result = collect_qualities(text_file_path)
        if result is None or len(result) != 3:
            preferred_qualities, midzone_qualities, out_qualities = [], [], []
        else:
            preferred_qualities, midzone_qualities, out_qualities = result
    except Exception as e:
        print(f"Error collecting qualities: {str(e)}")
        preferred_qualities, midzone_qualities, out_qualities = [], [], []
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Set up data sheet
    data_sheet = wb.active
    data_sheet.title = "MBTI Data"
    setup_data_sheet(data_sheet, info, mbti_dict, preferred_qualities, midzone_qualities, out_qualities)
    
    # Create dashboard sheet
    dashboard_sheet = wb.create_sheet(title="Dashboard")
    setup_dashboard_sheet(dashboard_sheet, info, mbti_dict)
    
    # Create insights sheet
    insights_sheet = wb.create_sheet(title="Insights")
    setup_insights_sheet(insights_sheet, info, preferred_qualities, midzone_qualities, out_qualities)
    
    # Save the workbook
    output_path = os.path.join(output_dir, output_filename)
    wb.save(output_path)
    
    return output_path


def setup_data_sheet(sheet, info, mbti_dict, preferred_qualities, midzone_qualities, out_qualities):
    """Set up the data sheet with MBTI information"""
    # Add title
    sheet['A1'] = f"MBTI Profile for {info['name']}"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:F1')
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Add basic information
    sheet['A3'] = "Name:"
    sheet['B3'] = info['name']
    sheet['A4'] = "Date:"
    sheet['B4'] = info['date']
    sheet['A5'] = "Type:"
    sheet['B5'] = info['type']
    
    # Add MBTI scores
    sheet['D3'] = "Extroversion:"
    sheet['E3'] = mbti_dict.get('E', 0)
    sheet['D4'] = "Introversion:"
    sheet['E4'] = mbti_dict.get('I', 0)
    sheet['D5'] = "Sensing:"
    sheet['E5'] = mbti_dict.get('S', 0)
    sheet['D6'] = "Intuition:"
    sheet['E6'] = mbti_dict.get('N', 0)
    sheet['D7'] = "Thinking:"
    sheet['E7'] = mbti_dict.get('T', 0)
    sheet['D8'] = "Feeling:"
    sheet['E8'] = mbti_dict.get('F', 0)
    sheet['D9'] = "Judging:"
    sheet['E9'] = mbti_dict.get('J', 0)
    sheet['D10'] = "Perceiving:"
    sheet['E10'] = mbti_dict.get('P', 0)
    
    # Format headers
    for cell in [sheet['A3'], sheet['D3']]:
        cell.font = Font(bold=True)
    
    # Add facets section
    sheet['A12'] = "Facets"
    sheet['A12'].font = Font(size=14, bold=True)
    sheet.merge_cells('A12:F12')
    sheet['A12'].alignment = Alignment(horizontal='center')
    
    # Add facet headers
    sheet['A14'] = "Facet"
    sheet['B14'] = "Preference"
    sheet['A14'].font = Font(bold=True)
    sheet['B14'].font = Font(bold=True)
    
    # Ensure qualities lists are not None
    preferred_qualities = preferred_qualities or []
    midzone_qualities = midzone_qualities or []
    out_qualities = out_qualities or []
    
    # Add facets data
    row = 15
    for facet in FACETS:
        sheet[f'A{row}'] = facet
        
        # Determine preference
        facet_lower = facet.lower()
        if facet_lower in [q.lower() for q in preferred_qualities]:
            sheet[f'B{row}'] = "IN-PREF"
            sheet[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif facet_lower in [q.lower() for q in midzone_qualities]:
            sheet[f'B{row}'] = "MIDZONE"
            sheet[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif facet_lower in [q.lower() for q in out_qualities]:
            sheet[f'B{row}'] = "OUT-OF-PREF"
            sheet[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            sheet[f'B{row}'] = "-"
        
        row += 1
    
    # Create a table for facets
    facets_table_ref = f"A14:B{row-1}"
    facets_table = Table(displayName="FacetsTable", ref=facets_table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    facets_table.tableStyleInfo = style
    sheet.add_table(facets_table)
    
    # Adjust column widths
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in sheet[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def setup_dashboard_sheet(sheet, info, mbti_dict):
    """Set up the dashboard sheet with charts and visualizations in a horizontal layout"""
    # Add title
    sheet['A1'] = f"MBTI Dashboard for {info['name']}"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:P1')
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Add type information with description
    sheet['A3'] = "Type:"
    sheet['B3'] = info['type']
    sheet['A3'].font = Font(bold=True)

    # Add type description
    sheet['A4'] = "Type Description:"
    sheet['A4'].font = Font(bold=True)
    sheet['B4'] = type_descriptions.get(info['type'], "Description not available")
    sheet.merge_cells('B4:P4')
    sheet['B4'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[4].height = 40

    # Create data for charts - place it in a less visible area
    sheet['A20'] = "Dimension"
    sheet['B20'] = "Score"
    sheet['A20'].font = Font(bold=True)
    sheet['B20'].font = Font(bold=True)

    # E-I dimension
    sheet['A21'] = "Extroversion"
    sheet['B21'] = mbti_dict.get('E', 0)
    sheet['A22'] = "Introversion"
    sheet['B22'] = mbti_dict.get('I', 0)

    # S-N dimension
    sheet['A23'] = "Sensing"
    sheet['B23'] = mbti_dict.get('S', 0)
    sheet['A24'] = "Intuition"
    sheet['B24'] = mbti_dict.get('N', 0)

    # T-F dimension
    sheet['A25'] = "Thinking"
    sheet['B25'] = mbti_dict.get('T', 0)
    sheet['A26'] = "Feeling"
    sheet['B26'] = mbti_dict.get('F', 0)

    # J-P dimension
    sheet['A27'] = "Judging"
    sheet['B27'] = mbti_dict.get('J', 0)
    sheet['A28'] = "Perceiving"
    sheet['B28'] = mbti_dict.get('P', 0)
    
    # Create summary table in a visible area
    sheet['A6'] = "MBTI Preferences Summary"
    sheet['A6'].font = Font(bold=True, size=12)
    sheet.merge_cells('A6:D6')
    sheet['A6'].alignment = Alignment(horizontal='center')
    
    # Headers
    sheet['A7'] = "Dimension"
    sheet['B7'] = "Score 1"
    sheet['C7'] = "Score 2"
    sheet['D7'] = "Preference"
    for cell in [sheet['A7'], sheet['B7'], sheet['C7'], sheet['D7']]:
        cell.font = Font(bold=True)
    
    # E-I row
    sheet['A8'] = "E-I"
    sheet['B8'] = mbti_dict.get('E', 0)
    sheet['C8'] = mbti_dict.get('I', 0)
    sheet['D8'] = "E" if mbti_dict.get('E', 0) > mbti_dict.get('I', 0) else "I"
    
    # S-N row
    sheet['A9'] = "S-N"
    sheet['B9'] = mbti_dict.get('S', 0)
    sheet['C9'] = mbti_dict.get('N', 0)
    sheet['D9'] = "S" if mbti_dict.get('S', 0) > mbti_dict.get('N', 0) else "N"
    
    # T-F row
    sheet['A10'] = "T-F"
    sheet['B10'] = mbti_dict.get('T', 0)
    sheet['C10'] = mbti_dict.get('F', 0)
    sheet['D10'] = "T" if mbti_dict.get('T', 0) > mbti_dict.get('F', 0) else "F"
    
    # J-P row
    sheet['A11'] = "J-P"
    sheet['B11'] = mbti_dict.get('J', 0)
    sheet['C11'] = mbti_dict.get('P', 0)
    sheet['D11'] = "J" if mbti_dict.get('J', 0) > mbti_dict.get('P', 0) else "P"
    
    # Create bar chart for all dimensions - place it in the middle
    bar_chart = BarChart()
    bar_chart.title = "MBTI Preferences"
    bar_chart.style = 10
    bar_chart.x_axis.title = "Dimension"
    bar_chart.y_axis.title = "Score"
    bar_chart.type = "col"  # Column chart
    bar_chart.grouping = "clustered"
    bar_chart.overlap = 0
    bar_chart.height = 15  # Adjust height
    bar_chart.width = 20   # Adjust width

    data = Reference(sheet, min_col=2, min_row=21, max_row=28, max_col=2)
    categories = Reference(sheet, min_col=1, min_row=21, max_row=28, max_col=1)

    bar_chart.add_data(data)
    bar_chart.set_categories(categories)

    sheet.add_chart(bar_chart, "F6")

    # Create pie charts for each dichotomy - arrange them horizontally
    # E-I
    create_dichotomy_pie_chart(sheet, "E-I Preference", 21, 22, "A13")

    # S-N
    create_dichotomy_pie_chart(sheet, "S-N Preference", 23, 24, "F13")

    # T-F
    create_dichotomy_pie_chart(sheet, "T-F Preference", 25, 26, "K13")

    # J-P
    create_dichotomy_pie_chart(sheet, "J-P Preference", 27, 28, "P13")

    # Adjust column widths
    for col in range(1, 17):  # Columns A through P
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 12  # Set a standard width
    
    # Hide the data rows used for charts
    for row in range(20, 29):
        sheet.row_dimensions[row].hidden = True


def setup_insights_sheet(sheet, info, preferred_qualities, midzone_qualities, out_qualities):
    """Set up the insights sheet with detailed information about the MBTI type and facets"""
    # Ensure qualities lists are not None
    preferred_qualities = preferred_qualities or []
    midzone_qualities = midzone_qualities or []
    out_qualities = out_qualities or []
    
    # Add title
    sheet['A1'] = f"MBTI Insights for {info['name']}"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:H1')
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Add type information
    sheet['A3'] = "Type:"
    sheet['B3'] = info.get('type', 'Unknown')
    sheet['A3'].font = Font(bold=True)

    # Add type insights
    sheet['A5'] = "Type Insights"
    sheet['A5'].font = Font(size=14, bold=True)
    sheet.merge_cells('A5:H5')
    sheet['A5'].alignment = Alignment(horizontal='center')

    # Type insights based on the four letters
    mbti_type = info.get('type', 'Unknown') or 'Unknown'  # Ensure mbti_type is not None

    # E/I insights
    sheet['A7'] = "Extraversion (E) vs. Introversion (I):"
    sheet['A7'].font = Font(bold=True)
    if mbti_type != 'Unknown' and 'E' in mbti_type:
        sheet['B7'] = "You gain energy from external interaction and tend to be more outgoing and sociable."
    elif mbti_type != 'Unknown' and 'I' in mbti_type:
        sheet['B7'] = "You gain energy from internal reflection and tend to be more reserved and thoughtful."
    else:
        sheet['B7'] = "Information not available."
    sheet.merge_cells('B7:H7')
    sheet['B7'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[7].height = 30

    # S/N insights
    sheet['A8'] = "Sensing (S) vs. Intuition (N):"
    sheet['A8'].font = Font(bold=True)
    if mbti_type != 'Unknown' and 'S' in mbti_type:
        sheet['B8'] = "You focus on concrete facts and details, preferring practical and realistic approaches."
    elif mbti_type != 'Unknown' and 'N' in mbti_type:
        sheet['B8'] = "You focus on patterns and possibilities, preferring abstract and theoretical approaches."
    else:
        sheet['B8'] = "Information not available."
    sheet.merge_cells('B8:H8')
    sheet['B8'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[8].height = 30

    # T/F insights
    sheet['A9'] = "Thinking (T) vs. Feeling (F):"
    sheet['A9'].font = Font(bold=True)
    if mbti_type != 'Unknown' and 'T' in mbti_type:
        sheet['B9'] = "You make decisions based on logical analysis and objective criteria."
    elif mbti_type != 'Unknown' and 'F' in mbti_type:
        sheet['B9'] = "You make decisions based on personal values and how they affect people."
    else:
        sheet['B9'] = "Information not available."
    sheet.merge_cells('B9:H9')
    sheet['B9'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[9].height = 30

    # J/P insights
    sheet['A10'] = "Judging (J) vs. Perceiving (P):"
    sheet['A10'].font = Font(bold=True)
    if mbti_type != 'Unknown' and 'J' in mbti_type:
        sheet['B10'] = "You prefer structure, planning, and organization in your life and work."
    elif mbti_type != 'Unknown' and 'P' in mbti_type:
        sheet['B10'] = "You prefer flexibility, spontaneity, and keeping options open in your life and work."
    else:
        sheet['B10'] = "Information not available."
    sheet.merge_cells('B10:H10')
    sheet['B10'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[10].height = 30

    # Add career insights
    sheet['A12'] = "Career Insights"
    sheet['A12'].font = Font(size=14, bold=True)
    sheet.merge_cells('A12:H12')
    sheet['A12'].alignment = Alignment(horizontal='center')
    # Career insights based on type

    sheet['A13'] = career_insights.get(mbti_type, "Career insights not available for this type.")
    sheet.merge_cells('A13:H15')
    sheet['A13'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[13].height = 60

    # Adjust column widths
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in sheet[column_letter]:
            if cell.value and isinstance(cell.value, str):
                max_length = max(max_length, len(cell.value))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def create_dichotomy_pie_chart(sheet, title, first_row, second_row, position):
    """Create a pie chart for a dichotomy"""
    pie_chart = PieChart()
    pie_chart.title = title
    pie_chart.style = 10
    
    data = Reference(sheet, min_col=2, min_row=first_row, max_row=second_row, max_col=2)
    labels = Reference(sheet, min_col=1, min_row=first_row, max_row=second_row, max_col=1)
    
    pie_chart.add_data(data)
    pie_chart.set_categories(labels)
    
    # Set slice colors - Note: Direct color setting is limited in openpyxl
    # We'll rely on the default colors which should be distinct enough
    
    sheet.add_chart(pie_chart, position)


if __name__ == "__main__":
    # For testing purposes
    input_pdf_path = r"F:\projects\MBTInfo\input\ADAM-POMERANTZ-267149-e4b6edb5-1a5f-ef11-bdfd-6045bd04b01a.pdf"
    output_dir = r"F:\projects\MBTInfo\output"
    output_filename = "Personal_MBTI_Report_Test.xlsx"
    
    if os.path.exists(input_pdf_path):
        output_path = generate_personal_report(input_pdf_path, output_dir, output_filename)
        print(f"Personal MBTI report generated at: {output_path}")
    else:
        print(f"Input file not found: {input_pdf_path}")
        print("Please provide a valid PDF file path for testing.")
