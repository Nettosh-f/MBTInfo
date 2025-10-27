from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .consts import (
    FACET_OCCURRENCE_COUNTS,
    FACET_TABLE_FACET_COL_END,
    FACET_TABLE_FACET_COL_START,
    FACET_TABLE_HEADERS,
    FACET_TABLE_ONE_TIME_COL_END,
    FACET_TABLE_ONE_TIME_COL_START,
    FACET_TABLE_ONE_TIME_COLOR,
    FACET_TABLE_ONE_TIME_INITIAL_COL,
    FACET_TABLE_STYLE,
    FACET_TABLE_TABLE_NAME,
    FACET_TABLE_THREE_TIMES_COL_END,
    FACET_TABLE_THREE_TIMES_COL_START,
    FACET_TABLE_THREE_TIMES_COLOR,
    FACET_TABLE_THREE_TIMES_INITIAL_COL,
    FACET_TABLE_TWO_TIMES_COL_END,
    FACET_TABLE_TWO_TIMES_COL_START,
    FACET_TABLE_TWO_TIMES_COLOR,
    FACET_TABLE_TWO_TIMES_INITIAL_COL,
    SHEET_NAME_FACET_TABLE,
    SHEET_NAME_MBTI_RESULTS,
)
from .formatting import adjust_column_widths


def create_facet_table(workbook):
    sheet_name = SHEET_NAME_FACET_TABLE

    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        # Remove the existing sheet
        workbook.remove(workbook[sheet_name])

    # Create a new sheet
    sheet = workbook.create_sheet(title=sheet_name)

    # Define headers from constants
    headers = FACET_TABLE_HEADERS

    # Define column colors
    three_times_color = PatternFill(
        start_color=FACET_TABLE_THREE_TIMES_COLOR,
        end_color=FACET_TABLE_THREE_TIMES_COLOR,
        fill_type="solid",
    )  # Light green
    two_times_color = PatternFill(
        start_color=FACET_TABLE_TWO_TIMES_COLOR,
        end_color=FACET_TABLE_TWO_TIMES_COLOR,
        fill_type="solid",
    )  # Light yellow
    one_time_color = PatternFill(
        start_color=FACET_TABLE_ONE_TIME_COLOR,
        end_color=FACET_TABLE_ONE_TIME_COLOR,
        fill_type="solid",
    )  # Light red

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

        # Apply colors to header cells based on column groups
        if (
            FACET_TABLE_THREE_TIMES_COL_START <= col <= FACET_TABLE_THREE_TIMES_COL_END
        ):  # Columns D, E, F (3 times)
            cell.fill = three_times_color
        elif (
            FACET_TABLE_TWO_TIMES_COL_START <= col <= FACET_TABLE_TWO_TIMES_COL_END
        ):  # Columns G, H, I, J, K (2 times)
            cell.fill = two_times_color
        elif (
            FACET_TABLE_ONE_TIME_COL_START <= col <= FACET_TABLE_ONE_TIME_COL_END
        ):  # Columns L through T (1 time)
            cell.fill = one_time_color

    # Get the 'MBTI Results' sheet
    mbti_results_sheet = workbook[SHEET_NAME_MBTI_RESULTS]
    max_row = mbti_results_sheet.max_row

    # Process each row (start from row 2 to skip headers)
    for row in range(2, max_row + 1):
        target_row = row  # Start from row 2 in the Facet Table

        # Copy Name, Date, and Type (columns 1-3)
        for col in range(1, 4):
            sheet.cell(
                row=target_row,
                column=col,
                value=mbti_results_sheet.cell(row=row, column=col).value,
            )

        # Get facets from AZ to BZ
        facets = [
            cell.value
            for cell in mbti_results_sheet[row][
                FACET_TABLE_FACET_COL_START:FACET_TABLE_FACET_COL_END
            ]
            if cell.value
        ]  # AZ is column 52

        # Count facet occurrences
        facet_counts = Counter(facets)

        # Write facets in order: 3 timers, 2 timers, 1 timers
        col = FACET_TABLE_THREE_TIMES_INITIAL_COL
        for count in FACET_OCCURRENCE_COUNTS:
            facets_with_count = [
                facet for facet, c in facet_counts.items() if c == count
            ]
            for facet in facets_with_count:
                sheet.cell(row=target_row, column=col, value=facet)
                col += 1

            # Move to the next group of columns
            if count == 3:
                col = FACET_TABLE_TWO_TIMES_INITIAL_COL  # Move to "Appearing 2 times" columns
            elif count == 2:
                col = FACET_TABLE_ONE_TIME_INITIAL_COL  # Move to "Appearing 1 time" columns

    last_row = sheet.max_row
    last_col = sheet.max_column
    last_col_letter = get_column_letter(last_col)
    table_ref = f"A1:{last_col_letter}{last_row}"

    # Remove old table if exists
    if FACET_TABLE_TABLE_NAME in sheet.tables:
        del sheet.tables[FACET_TABLE_TABLE_NAME]

    table = Table(displayName=FACET_TABLE_TABLE_NAME, ref=table_ref)
    style = TableStyleInfo(
        name=FACET_TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)
    # Adjust column widths
    adjust_column_widths(sheet)

    return workbook


if __name__ == "__main__":
    # For testing purposes
    workbook_path = r"F:\projects\MBTInfo\output\MBTI_Results.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    create_facet_table(workbook)
    workbook.save(workbook_path)
