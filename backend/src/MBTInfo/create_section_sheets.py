import os

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from .consts import (
    SECTION_SHEET_BASE_HEADERS,
    SECTION_SHEET_COLUMN_PADDING,
    SECTION_SHEET_MAX_FACETS,
    SECTION_SHEET_NAMES,
    SECTION_SHEET_STYLE,
    SECTION_SHEET_TABLE_END_COLUMN,
    SECTION_SHEET_TABLE_START,
    SECTION_SHEET_TEXT_FILE_SUFFIX,
    SECTION_SHEET_TITLE_FONT_SIZE,
)
from .utils import (
    check_communication,
    check_managing_change,
    check_managing_conflict,
    get_all_info,
)


def create_section_sheets(input_directory, workbook):
    sections = SECTION_SHEET_NAMES
    check_functions = [
        check_communication,
        check_managing_change,
        check_managing_conflict,
    ]

    for section, check_function in zip(sections, check_functions):
        sheet = workbook.create_sheet(title=section)
        _setup_sheet(sheet, section)

        # Start data from row 4 (rows 1-3 used for title/headers)
        row = 4
        for file in os.listdir(input_directory):
            if file.endswith(SECTION_SHEET_TEXT_FILE_SUFFIX):
                text_path = os.path.join(input_directory, file)
                info = get_all_info(text_path)
                facets = check_function(text_path)
                _add_data_row(sheet, row, info, facets)
                row += 1

        _create_table(sheet, row - 1)  # Create table after all data is added

    return workbook


def _setup_sheet(sheet, title):
    # Add title
    sheet["A1"] = title
    sheet["A1"].font = Font(size=SECTION_SHEET_TITLE_FONT_SIZE, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")

    # Add headers (row 3)
    headers = SECTION_SHEET_BASE_HEADERS + [
        f"Facet {i}" for i in range(1, SECTION_SHEET_MAX_FACETS + 1)
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)


def _add_data_row(sheet, row, info, facets):
    data = (
        [info["name"], info["date"], info["type"]]
        + facets
        + [""] * (SECTION_SHEET_MAX_FACETS - len(facets))
    )
    for col, value in enumerate(data, start=1):
        sheet.cell(row=row, column=col, value=value)


def _create_table(sheet, last_row):
    table_ref = (
        f"{SECTION_SHEET_TABLE_START}:{SECTION_SHEET_TABLE_END_COLUMN}{last_row}"
    )
    tab = Table(displayName=f"{sheet.title.replace(' ', '')}Table", ref=table_ref)
    style = TableStyleInfo(
        name=SECTION_SHEET_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    # Adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        sheet.column_dimensions[column_cells[0].column_letter].width = (
            length + SECTION_SHEET_COLUMN_PADDING
        )
