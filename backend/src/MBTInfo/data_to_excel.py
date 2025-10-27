import os

import openpyxl as xl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .consts import (
    DATA_EXCEL_BASE_HEADERS,
    DATA_EXCEL_COLUMN_PADDING,
    DATA_EXCEL_DEFAULT_WIDTH,
    DATA_EXCEL_FREEZE_PANES,
    DATA_EXCEL_SECTION_COLUMNS,
    DATA_EXCEL_TABLE_NAME,
    DATA_EXCEL_TABLE_START,
    DATA_EXCEL_TABLE_STYLE,
    FACETS,
    SECTION_SHEET_MAX_FACETS,
    SECTION_SHEET_NAMES,
)
from .utils import (
    check_communication,
    check_managing_change,
    check_managing_conflict,
    collect_qualities,
    convert_scores_to_mbti_dict,
    find_and_parse_mbti_scores,
    get_all_info,
)


def process_pdf_to_xl(text_path, output_dir, result_sheet_name, output_filename):
    info = get_all_info(text_path)
    qualities = find_and_parse_mbti_scores(text_path)
    mbti_dict = convert_scores_to_mbti_dict(qualities)
    preferred_qualities, midzone_qualities, out_qualities = collect_qualities(text_path)

    # Get communication, change management, and conflict management facets
    communication_facets = check_communication(text_path)
    change_facets = check_managing_change(text_path)
    conflict_facets = check_managing_conflict(text_path)

    # Convert all qualities to lowercase for case-insensitive comparison
    preferred_qualities = [q.lower() for q in preferred_qualities]
    midzone_qualities = [q.lower() for q in midzone_qualities]
    out_qualities = [q.lower() for q in out_qualities]

    output_path = os.path.join(output_dir, output_filename)

    headers = DATA_EXCEL_BASE_HEADERS.copy()
    headers.extend(FACETS)  # Add all facets to the headers

    sections = SECTION_SHEET_NAMES
    headers.extend(sections)  # Add section headers without empty cells

    # Calculate the actual number of columns needed
    actual_columns = len(headers)
    # Add extra columns for the section data
    actual_columns += DATA_EXCEL_SECTION_COLUMNS * len(sections)

    if os.path.exists(output_path):
        workbook = xl.load_workbook(filename=str(output_path))
        if result_sheet_name in workbook.sheetnames:
            sheet = workbook[result_sheet_name]

            # Find the last non-empty row
            last_row = sheet.max_row
            while last_row > 1 and all(cell.value is None for cell in sheet[last_row]):
                last_row -= 1
            last_row += 1  # Move to the next empty row
        else:
            sheet = workbook.create_sheet(result_sheet_name)
            _setup_headers(sheet, headers, sections)
            last_row = 2  # Start data from row 2 (row 1 is headers)
    else:
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = result_sheet_name
        _setup_headers(sheet, headers, sections)
        last_row = 2  # Start data from row 2 (row 1 is headers)

    data = [info["name"], info["date"], info["type"]] + list(mbti_dict.values())

    # Add values for facets (exclude base headers and section headers)
    # Slice to get only facet headers (after base headers, before section headers)
    num_base_headers = len(DATA_EXCEL_BASE_HEADERS)
    num_sections = len(sections)
    facet_headers = (
        headers[num_base_headers:-num_sections]
        if num_sections > 0
        else headers[num_base_headers:]
    )

    for header in facet_headers:
        header_lower = header.lower()
        if header_lower in preferred_qualities:
            data.append("IN-PREF")
        elif header_lower in midzone_qualities:
            data.append("MIDZONE")
        elif header_lower in out_qualities:
            data.append("OUT-OF-PREF")
        else:
            data.append("-")

    # Add values for communication, change management, and conflict management
    data.extend(
        communication_facets
        + [""] * (SECTION_SHEET_MAX_FACETS - len(communication_facets))
    )
    data.extend(change_facets + [""] * (SECTION_SHEET_MAX_FACETS - len(change_facets)))
    data.extend(
        conflict_facets + [""] * (SECTION_SHEET_MAX_FACETS - len(conflict_facets))
    )

    # Append data to the last empty row
    for col, value in enumerate(data, start=1):
        sheet.cell(row=last_row, column=col, value=value if value else "")

    # Update or create the table
    last_row = sheet.max_row
    last_col = sheet.max_column
    last_col_letter = get_column_letter(last_col)
    table_range = f"{DATA_EXCEL_TABLE_START}:{last_col_letter}{last_row}"

    # Remove old table if exists
    if DATA_EXCEL_TABLE_NAME in sheet.tables:
        del sheet.tables[DATA_EXCEL_TABLE_NAME]

    # Create a new table with the correct range and style
    tab = Table(displayName=DATA_EXCEL_TABLE_NAME, ref=table_range)
    style = TableStyleInfo(
        name=DATA_EXCEL_TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    workbook.save(filename=str(output_path))
    return output_path


def _unmerge_first_row(sheet):
    for merge_range in list(sheet.merged_cells.ranges):
        if merge_range.min_row == 1:
            sheet.unmerge_cells(str(merge_range))


def _setup_headers(sheet, headers, sections):
    col = 1
    for header in headers:
        cell = sheet.cell(row=1, column=col)
        # Assign placeholder for empty cells
        cell.value = header if header not in sections else header
        cell.font = Font(bold=True)

        if header in sections:
            cell.alignment = Alignment(horizontal="center")
            # Add named cells after the section header
            for i in range(DATA_EXCEL_SECTION_COLUMNS):
                col += 1
                sheet.cell(row=1, column=col).value = f"{header}-{i + 1}"
        col += 1
    # Adjust column widths to fit content
    for column_cells in sheet.columns:
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value]
        if lengths:
            max_length = max(lengths)
            sheet.column_dimensions[column_cells[0].column_letter].width = (
                max_length + DATA_EXCEL_COLUMN_PADDING
            )
        else:
            # Set a default width for empty columns
            sheet.column_dimensions[column_cells[0].column_letter].width = (
                DATA_EXCEL_DEFAULT_WIDTH
            )

    # Freeze the first row
    sheet.freeze_panes = DATA_EXCEL_FREEZE_PANES


if __name__ == "__main__":
    pdf_path = r"F:\projects\MBTInfo\output\textfiles\nir-bensinai-MBTI_text.txt"
    output_dir = r"F:\projects\MBTInfo\output"
    output_filename = "MBTI_Results_test.xlsx"
    result_sheet_name = "MBTI Results"
    output_path = process_pdf_to_xl(
        pdf_path, output_dir, result_sheet_name, output_filename
    )
    print(f"MBTI results appended to {output_path}")
