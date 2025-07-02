import openpyxl
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
import openpyxl.chart.text
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
# local import
from consts import mbti_colors, MBTI_TYPES


def create_distribution_charts(workbook):
    # Create or get the data sheet
    if 'Data' in workbook.sheetnames:
        data_sheet = workbook['Data']
    else:
        data_sheet = workbook.create_sheet(title="Data")

    # Create or get the dashboard sheet for charts
    if 'Dashboard' in workbook.sheetnames:
        chart_sheet = workbook['Dashboard']
    else:
        chart_sheet = workbook.create_sheet(title="Dashboard")

    # Clear any existing charts
    chart_sheet._charts = []

    # Add a title to the dashboard
    chart_sheet['D1'] = "MBTI Distribution Dashboard"
    chart_sheet['D1'].font = Font(bold=True, size=16)

    # Prepare data in the data sheet
    prepare_main_distribution_data(data_sheet)
    prepare_dichotomy_data(data_sheet)
    prepare_external_internal_data(data_sheet)
    prepare_dominant_function_data(data_sheet)

    # Create charts in the dashboard sheet
    create_main_distribution_chart(data_sheet, chart_sheet)
    create_dichotomy_charts(data_sheet, chart_sheet)
    create_external_internal_charts(data_sheet, chart_sheet)
    create_facet_bar_charts(data_sheet, chart_sheet)
    create_dominant_chart(data_sheet, chart_sheet)
    # create_legend_for_facet_graphs(chart_sheet)

    # Adjust column widths for data sheet
    adjust_column_widths(data_sheet)

    # Hide gridlines in the dashboard for a cleaner look
    chart_sheet.sheet_view.showGridLines = False


def prepare_main_distribution_data(data_sheet):
    # Write headers
    data_sheet['A1'] = "MBTI Type Data"
    data_sheet['A1'].font = Font(bold=True)

    # Write MBTI types and count formulas
    for row, mbti_type in enumerate(MBTI_TYPES, start=2):
        data_sheet[f'A{row}'] = mbti_type
        data_sheet[f'B{row}'] = f"=COUNTIF(Table1[Type],A{row})"
        if mbti_type in mbti_colors:
            data_sheet[f'A{row}'].fill = PatternFill(start_color=mbti_colors[mbti_type],
                                                     end_color=mbti_colors[mbti_type],
                                                     fill_type="solid")


def prepare_dichotomy_data(data_sheet):
    data_sheet['D1'] = "Dichotomies Data"
    data_sheet['D1'].font = Font(bold=True)
    data_sheet['E2'] = "Count"
    data_sheet['E2'].font = Font(bold=True)

    # Define dichotomy colors - using professional color scheme
    dichotomy_colors = {
        "Extroversion": "4472C4",  # Blue
        "Introversion": "ED7D31",  # Orange
        "Sensing": "70AD47",  # Green
        "Intuition": "FFC000",  # Gold
        "Thinking": "5B9BD5",  # Light Blue
        "Feeling": "A5A5A5",  # Gray
        "Judging": "9966FF",  # Purple
        "Perceiving": "4BACC6"  # Teal
    }

    # Energy source - E/I
    data_sheet['D2'] = "Energy Orientation (E/I)"
    data_sheet['D2'].font = Font(bold=True)
    data_sheet['D3'] = "Extroversion"
    data_sheet['D4'] = "Introversion"
    data_sheet['E3'] = "=COUNTIF(Table1[Type], \"*E*\")"
    data_sheet['E4'] = "=COUNTIF(Table1[Type], \"*I*\")"
    data_sheet['D3'].fill = PatternFill(start_color=dichotomy_colors["Extroversion"],
                                        end_color=dichotomy_colors["Extroversion"], fill_type="solid")
    data_sheet['D4'].fill = PatternFill(start_color=dichotomy_colors["Introversion"],
                                        end_color=dichotomy_colors["Introversion"], fill_type="solid")

    # Information - S/N
    data_sheet['D6'] = "Information Gathering (S/N)"
    data_sheet['D6'].font = Font(bold=True)
    data_sheet['D7'] = "Sensing"
    data_sheet['D8'] = "Intuition"
    data_sheet['E7'] = "=COUNTIF(Table1[Type], \"*S*\")"
    data_sheet['E8'] = "=COUNTIF(Table1[Type], \"*N*\")"
    data_sheet['D7'].fill = PatternFill(start_color=dichotomy_colors["Sensing"],
                                        end_color=dichotomy_colors["Sensing"], fill_type="solid")
    data_sheet['D8'].fill = PatternFill(start_color=dichotomy_colors["Intuition"],
                                        end_color=dichotomy_colors["Intuition"], fill_type="solid")

    # Decisions - T/F
    data_sheet['D10'] = "Decision Making (T/F)"
    data_sheet['D10'].font = Font(bold=True)
    data_sheet['D11'] = "Thinking"
    data_sheet['D12'] = "Feeling"
    data_sheet['E11'] = "=COUNTIF(Table1[Type], \"*T*\")"
    data_sheet['E12'] = "=COUNTIF(Table1[Type], \"*F*\")"
    data_sheet['D11'].fill = PatternFill(start_color=dichotomy_colors["Thinking"],
                                        end_color=dichotomy_colors["Thinking"], fill_type="solid")
    data_sheet['D12'].fill = PatternFill(start_color=dichotomy_colors["Feeling"],
                                         end_color=dichotomy_colors["Feeling"], fill_type="solid")

    # Lifestyle - J/P
    data_sheet['D14'] = "Lifestyle & Structure (J/P)"
    data_sheet['D14'].font = Font(bold=True)
    data_sheet['D15'] = "Judging"
    data_sheet['D16'] = "Perceiving"
    data_sheet['E15'] = "=COUNTIF(Table1[Type], \"*J*\")"
    data_sheet['E16'] = "=COUNTIF(Table1[Type], \"*P*\")"
    data_sheet['D15'].fill = PatternFill(start_color=dichotomy_colors["Judging"],
                                         end_color=dichotomy_colors["Judging"], fill_type="solid")
    data_sheet['D16'].fill = PatternFill(start_color=dichotomy_colors["Perceiving"],
                                         end_color=dichotomy_colors["Perceiving"], fill_type="solid")

    # calculate dichotomy preference
    data_sheet['D19'] = "Dichotomy Preference data"
    data_sheet['D19'].font = Font(bold=True)

    # titles Initiating-Receiving
    data_sheet['D20'] = "Initiating-Receiving"
    data_sheet['D20'].font = Font(bold=True)
    data_sheet['D21'] = "Initiating"
    data_sheet['D22'] = "Receiving"
    data_sheet['D23'] = "MIDZONE"

    # data for Initiating-Receiving
    data_sheet['E21'] = '=COUNTIF(Table1[Initiating],"=IN-PREF")+COUNTIF(Table1[Initiating],"=OUT-OF-PREF")'
    data_sheet['E22'] = '=COUNTIF(Table1[Receiving],"=IN-PREF")+COUNTIF(Table1[Receiving],"=OUT-OF-PREF")'
    data_sheet['E23'] = '=COUNTIF(Table1[Initiating],"=MIDZONE")'

    # titles Initiating-Receiving
    data_sheet['D26'] = "Expressive-Contained"
    data_sheet['D26'].font = Font(bold=True)
    data_sheet['D27'] = "Expressive"
    data_sheet['D28'] = "Contained"
    data_sheet['D29'] = "MIDZONE"

    # data for Initiating-Receiving
    data_sheet['E26'] = '=COUNTIF(Table1[Expressive],"=IN-PREF")+COUNTIF(Table1[Expressive],"=OUT-OF-PREF")'
    data_sheet['E27'] = '=COUNTIF(Table1[Contained],"=IN-PREF")+COUNTIF(Table1[Contained],"=OUT-OF-PREF")'
    data_sheet['E28'] = '=COUNTIF(Table1[Contained],"=MIDZONE")'

    # titles Gregarious-Intimate
    data_sheet['D30'] = "Gregarious-Intimate"
    data_sheet['D30'].font = Font(bold=True)
    data_sheet['D31'] = "Gregarious"
    data_sheet['D32'] = "Intimate"
    data_sheet['D33'] = "MIDZONE"

    # data for Gregarious-Intimate
    data_sheet['E31'] = '=COUNTIF(Table1[Gregarious],"=IN-PREF")+COUNTIF(Table1[Gregarious],"=OUT-OF-PREF")'
    data_sheet['E32'] = '=COUNTIF(Table1[Intimate],"=IN-PREF")+COUNTIF(Table1[Intimate],"=OUT-OF-PREF")'
    data_sheet['E33'] = '=COUNTIF(Table1[Intimate],"=MIDZONE")'

    # titles Active-Reflective
    data_sheet['D36'] = "Active-Reflective"
    data_sheet['D36'].font = Font(bold=True)
    data_sheet['D37'] = "Active"
    data_sheet['D38'] = "Reflective"
    data_sheet['D39'] = "MIDZONE"

    # data for Active-Reflective
    data_sheet['E37'] = '=COUNTIF(Table1[Active],"=IN-PREF")+COUNTIF(Table1[Active],"=OUT-OF-PREF")'
    data_sheet['E38'] = '=COUNTIF(Table1[Reflective],"=IN-PREF")+COUNTIF(Table1[Reflective],"=OUT-OF-PREF")'
    data_sheet['E39'] = '=COUNTIF(Table1[Reflective],"=MIDZONE")'

    # titles Enthusiastic-Quiet
    data_sheet['D42'] = "Enthusiastic-Quiet"
    data_sheet['D42'].font = Font(bold=True)
    data_sheet['D43'] = "Enthusiastic"
    data_sheet['D44'] = "Quiet"
    data_sheet['D45'] = "MIDZONE"

    # data for Enthusiastic-Quiet
    data_sheet['E43'] = '=COUNTIF(Table1[Enthusiastic],"=IN-PREF")+COUNTIF(Table1[Enthusiastic],"=OUT-OF-PREF")'
    data_sheet['E44'] = '=COUNTIF(Table1[Quiet],"=IN-PREF")+COUNTIF(Table1[Quiet],"=OUT-OF-PREF")'
    data_sheet['E45'] = '=COUNTIF(Table1[Quiet],"=MIDZONE")'

    # titles Concrete-Abstract
    data_sheet['D48'] = "Concrete-Abstract"
    data_sheet['D48'].font = Font(bold=True)
    data_sheet['D49'] = "Concrete"
    data_sheet['D50'] = "Abstract"
    data_sheet['D51'] = "MIDZONE"

    # data for Concrete-Abstract
    data_sheet['E49'] = '=COUNTIF(Table1[Concrete],"=IN-PREF")+COUNTIF(Table1[Concrete],"=OUT-OF-PREF")'
    data_sheet['E50'] = '=COUNTIF(Table1[Abstract],"=IN-PREF")+COUNTIF(Table1[Abstract],"=OUT-OF-PREF")'
    data_sheet['E51'] = '=COUNTIF(Table1[Abstract],"=MIDZONE")'

    # titles Realistic-Imaginative
    data_sheet['D54'] = "Realistic-Imaginative"
    data_sheet['D54'].font = Font(bold=True)
    data_sheet['D55'] = "Realistic"
    data_sheet['D56'] = "Imaginative"
    data_sheet['D57'] = "MIDZONE"

    # data for Realistic-Imaginative
    data_sheet['E55'] = '=COUNTIF(Table1[Realistic],"=IN-PREF")+COUNTIF(Table1[Realistic],"=OUT-OF-PREF")'
    data_sheet['E56'] = '=COUNTIF(Table1[Imaginative],"=IN-PREF")+COUNTIF(Table1[Imaginative],"=OUT-OF-PREF")'
    data_sheet['E57'] = '=COUNTIF(Table1[Imaginative],"=MIDZONE")'

    # titles Practical-Conceptual
    data_sheet['D60'] = "Practical-Conceptual"
    data_sheet['D60'].font = Font(bold=True)
    data_sheet['D61'] = "Practical"
    data_sheet['D62'] = "Conceptual"
    data_sheet['D63'] = "MIDZONE (Practical-Conceptual)"
    # data for Practical-Conceptual
    data_sheet['E61'] = '=COUNTIF(Table1[Practical],"=IN-PREF")+COUNTIF(Table1[Practical],"=OUT-OF-PREF")'
    data_sheet['E62'] = '=COUNTIF(Table1[Conceptual],"=IN-PREF")+COUNTIF(Table1[Conceptual],"=OUT-OF-PREF")'
    data_sheet['E63'] = '=COUNTIF'

    # titles Experiential-Theoretical
    data_sheet['D66'] = "Experiential-Theoretical"
    data_sheet['D66'].font = Font(bold=True)
    data_sheet['D67'] = "Experiential"
    data_sheet['D68'] = "Theoretical"
    data_sheet['D69'] = "MIDZONE"
    # data for Experiential-Theoretical
    data_sheet['E67'] = '=COUNTIF(Table1[Experiential],"=IN-PREF")+COUNTIF(Table1[Experiential],"=OUT-OF-PREF")'
    data_sheet['E68'] = '=COUNTIF(Table1[Theoretical],"=IN-PREF")+COUNTIF(Table1[Theoretical],"=OUT-OF-PREF")'
    data_sheet['E69'] = '=COUNTIF(Table1[Theoretical],"=MIDZONE")'

    # titles Traditional-Original
    data_sheet['D72'] = "Traditional-Original"
    data_sheet['D72'].font = Font(bold=True)
    data_sheet['D73'] = "Traditional"
    data_sheet['D74'] = "Original"
    data_sheet['D75'] = "MIDZONE"
    # data for Traditional-Original
    data_sheet['E73'] = '=COUNTIF(Table1[Traditional],"=IN-PREF")+COUNTIF(Table1[Traditional],"=OUT-OF-PREF")'
    data_sheet['E74'] = '=COUNTIF(Table1[Original],"=IN-PREF")+COUNTIF(Table1[Original],"=OUT-OF-PREF")'
    data_sheet['E75'] = '=COUNTIF(Table1[Original],"=MIDZONE")'

    # titles Logical-Empathetic
    data_sheet['D78'] = "Logical-Empathetic"
    data_sheet['D78'].font = Font(bold=True)
    data_sheet['D79'] = "Logical"
    data_sheet['D80'] = "Empathetic"
    data_sheet['D81'] = "MIDZONE"
    # data for Logical-Empathetic
    data_sheet['E79'] = '=COUNTIF(Table1[Logical],"=IN-PREF")+COUNTIF(Table1[Logical],"=OUT-OF-PREF")'
    data_sheet['E80'] = '=COUNTIF(Table1[Empathetic],"=IN-PREF")+COUNTIF(Table1[Empathetic],"=OUT-OF-PREF")'
    data_sheet['E81'] = '=COUNTIF(Table1[Empathetic],"=MIDZONE")'

    # titles Reasonable-Compassionate
    data_sheet['D84'] = "Reasonable-Compassionate"
    data_sheet['D84'].font = Font(bold=True)
    data_sheet['D85'] = "Reasonable"
    data_sheet['D86'] = "Compassionate"
    data_sheet['D87'] = "MIDZONE"
    # data for Reasonable-Compassionate
    data_sheet['E85'] = '=COUNTIF(Table1[Reasonable],"=IN-PREF")+COUNTIF(Table1[Reasonable],"=OUT-OF-PREF")'
    data_sheet['E86'] = '=COUNTIF(Table1[Compassionate],"=IN-PREF")+COUNTIF(Table1[Compassionate],"=OUT-OF-PREF")'
    data_sheet['E87'] = '=COUNTIF(Table1[Compassionate],"=MIDZONE")'

    # titles Questioning-Accommodating
    data_sheet['D90'] = "Questioning-Accommodating"
    data_sheet['D90'].font = Font(bold=True)
    data_sheet['D91'] = "Questioning"
    data_sheet['D92'] = "Accommodating"
    data_sheet['D93'] = "MIDZONE"
    # data for Questioning-Accommodating
    data_sheet['E91'] = '=COUNTIF(Table1[Questioning],"=IN-PREF")+COUNTIF(Table1[Questioning],"=OUT-OF-PREF")'
    data_sheet['E92'] = '=COUNTIF(Table1[Accommodating],"=IN-PREF")+COUNTIF(Table1[Accommodating],"=OUT-OF-PREF")'
    data_sheet['E93'] = '=COUNTIF(Table1[Accommodating],"=MIDZONE")'

    # titles Critical-Accepting
    data_sheet['D96'] = "Critical-Accepting"
    data_sheet['D96'].font = Font(bold=True)
    data_sheet['D97'] = "Critical"
    data_sheet['D98'] = "Accepting"
    data_sheet['D99'] = "MIDZONE"
    # data for Critical-Accepting
    data_sheet['E97'] = '=COUNTIF(Table1[Critical],"=IN-PREF")+COUNTIF(Table1[Critical],"=OUT-OF-PREF")'
    data_sheet['E98'] = '=COUNTIF(Table1[Accepting],"=IN-PREF")+COUNTIF(Table1[Accepting],"=OUT-OF-PREF")'
    data_sheet['E99'] = '=COUNTIF(Table1[Accepting],"=MIDZONE")'

    # titles Tough-Tender
    data_sheet['D102'] = "Tough-Tender"
    data_sheet['D102'].font = Font(bold=True)
    data_sheet['D103'] = "Tough"
    data_sheet['D104'] = "Tender"
    data_sheet['D105'] = "MIDZONE"
    # data for Tough-Tender
    data_sheet['E103'] = '=COUNTIF(Table1[Tough],"=IN-PREF")+COUNTIF(Table1[Tough],"=OUT-OF-PREF")'
    data_sheet['E104'] = '=COUNTIF(Table1[Tender],"=IN-PREF")+COUNTIF(Table1[Tender],"=OUT-OF-PREF")'
    data_sheet['E105'] = '=COUNTIF(Table1[Tender],"=MIDZONE")'

    # titles Systematic-Casual
    data_sheet['D108'] = "Systematic-Casual"
    data_sheet['D108'].font = Font(bold=True)
    data_sheet['D109'] = "Systematic"
    data_sheet['D110'] = "Casual"
    data_sheet['D111'] = "MIDZONE"
    # data for Systematic-Casual
    data_sheet['E109'] = '=COUNTIF(Table1[Systematic],"=IN-PREF")+COUNTIF(Table1[Systematic],"=OUT-OF-PREF")'
    data_sheet['E110'] = '=COUNTIF(Table1[Casual],"=IN-PREF")+COUNTIF(Table1[Casual],"=OUT-OF-PREF")'
    data_sheet['E111'] = '=COUNTIF(Table1[Casual],"=MIDZONE")'

    # titles Planful-Open-Ended
    data_sheet['D114'] = "Planful-Open-Ended"
    data_sheet['D114'].font = Font(bold=True)
    data_sheet['D115'] = "Planful"
    data_sheet['D116'] = "Open-Ended"
    data_sheet['D117'] = "MIDZONE"
    # data for Planful-Open-Ended
    data_sheet['E115'] = '=COUNTIF(Table1[Planful],"=IN-PREF")+COUNTIF(Table1[Planful],"=OUT-OF-PREF")'
    data_sheet['E116'] = '=COUNTIF(Table1[Open-Ended],"=IN-PREF")+COUNTIF(Table1[Open-Ended],"=OUT-OF-PREF")'
    data_sheet['E117'] = '=COUNTIF(Table1[Open-Ended],"=MIDZONE")'

    # titles Early Starting-Pressure-Prompted
    data_sheet['D120'] = "Early Starting-Pressure-Prompted"
    data_sheet['D120'].font = Font(bold=True)
    data_sheet['D121'] = "Early Starting"
    data_sheet['D122'] = "Pressure-Prompted"
    data_sheet['D123'] = "MIDZONE"
    # data for Early Starting-Pressure-Prompted
    data_sheet['E121'] = '=COUNTIF(Table1[Early Starting],"=IN-PREF")+COUNTIF(Table1[Early Starting],"=OUT-OF-PREF")'
    data_sheet['E122'] = '=COUNTIF(Table1[Pressure-Prompted],"=IN-PREF")+COUNTIF(Table1[Pressure-Prompted],"=OUT-OF-PREF")'
    data_sheet['E123'] = '=COUNTIF(Table1[Early Starting],"=MIDZONE")'

    # titles Scheduled-Spontaneous
    data_sheet['D126'] = "Scheduled-Spontaneous"
    data_sheet['D126'].font = Font(bold=True)
    data_sheet['D127'] = "Schedule"
    data_sheet['D128'] = "Spontaneous"
    data_sheet['D129'] = "MIDZONE"
    # data for Scheduled-Spontaneous
    data_sheet['E127'] = '=COUNTIF(Table1[Scheduled],"=IN-PREF")+COUNTIF(Table1[Scheduled],"=OUT-OF-PREF")'
    data_sheet['E128'] = '=COUNTIF(Table1[Spontaneous],"=IN-PREF")+COUNTIF(Table1[Spontaneous],"=OUT-OF-PREF")'
    data_sheet['E129'] = '=COUNTIF(Table1[Spontaneous],"=MIDZONE")'

    # titles Methodical-Emergent
    data_sheet['D132'] = "Methodical-Emergent"
    data_sheet['D132'].font = Font(bold=True)
    data_sheet['D133'] = "Methodical"
    data_sheet['D134'] = "Emergent"
    data_sheet['D135'] = "MIDZONE"
    # data for Methodical-Emergent
    data_sheet['E133'] = '=COUNTIF(Table1[Methodical],"=IN-PREF")+COUNTIF(Table1[Methodical],"=OUT-OF-PREF")'
    data_sheet['E134'] = '=COUNTIF(Table1[Emergent],"=IN-PREF")+COUNTIF(Table1[Emergent],"=OUT-OF-PREF")'
    data_sheet['E135'] = '=COUNTIF(Table1[Emergent],"=MIDZONE")'


def prepare_facet_legend(chart_sheet):
    # add white text rule to the black background for text readability

    # add text
    chart_sheet['V30'] = "Facets Legend"
    chart_sheet['V30'].font = Font(bold=True)

    chart_sheet['X31'] = "1ST facet - In-Preference"
    chart_sheet['X32'] = "2ND facet - In-Preference"
    chart_sheet['X33'] = "1ST facet - Out-of-Preference"
    chart_sheet['X34'] = "2ND facet - Out-of-Preference"
    chart_sheet['X35'] = "MIDZONE"


def prepare_external_internal_data(data_sheet):
    Internal_list = ["ST", "SF", "NF", "NT"]
    External_list = ["IJ", "IP", "EJ", "EP"]

    data_sheet['G1'] = "Internal Analysis"
    data_sheet['J1'] = "External Analysis"
    data_sheet['G1'].font = Font(bold=True)
    data_sheet['J1'].font = Font(bold=True)
    data_sheet['G2'] = "Internal"
    data_sheet['J2'] = "External"
    data_sheet['H2'] = "Count"
    data_sheet['K2'] = "Count"
    data_sheet['H3'] = '=COUNTIF(Table1[Type], "*ST*")'
    data_sheet['H4'] = '=COUNTIF(Table1[Type], "*SF*")'
    data_sheet['H5'] = '=COUNTIF(Table1[Type], "*NF*")'
    data_sheet['H6'] = '=COUNTIF(Table1[Type], "*NT*")'
    data_sheet['K3'] = '=COUNTIF(Table1[Type], "I*J")'
    data_sheet['K4'] = '=COUNTIF(Table1[Type], "I*P")'
    data_sheet['K5'] = '=COUNTIF(Table1[Type], "E*J")'
    data_sheet['K6'] = '=COUNTIF(Table1[Type], "E*P")'

    # Insert Internal_list at G3->G6
    for i, internal_type in enumerate(Internal_list):
        data_sheet[f'G{3 + i}'] = internal_type

    # Insert External_list at J3->J6
    for i, external_type in enumerate(External_list):
        data_sheet[f'J{3 + i}'] = external_type


def prepare_dominant_function_data(data_sheet):


    data_sheet['G9'] = "Dominant Function Data"
    data_sheet['G9'].font = Font(bold=True)
    data_sheet['G10'] = "Dominant Function"
    data_sheet['H10'] = "Count"
    data_sheet['G11'] = "Te"
    data_sheet['H11'] = '=SUM(B17,B14)'
    data_sheet['G12'] = "Ti"
    data_sheet['H12'] = '=SUM(B9,B6)'
    data_sheet['G13'] = "Ne"
    data_sheet['H13'] = '=SUM(B12,B13)'
    data_sheet['G14'] = "Ni"
    data_sheet['H14'] = '=SUM(B4,B5)'
    data_sheet['G15'] = "Fe"
    data_sheet['H15'] = '=SUM(B16,B15)'
    data_sheet['G16'] = "Fi"
    data_sheet['H16'] = '=SUM(B8,B7)'
    data_sheet['G17'] = "Se"
    data_sheet['H17'] = '=SUM(B11,B10)'
    data_sheet['G18'] = "Si"
    data_sheet['H18'] = '=SUM(B3,B2)'


def create_main_distribution_chart(data_sheet, chart_sheet):
    # Create pie chart
    main_chart = PieChart()
    main_chart.title = "Distribution by Type"

    # Set data references
    labels = Reference(data_sheet, min_col=1, min_row=2, max_row=17)
    data = Reference(data_sheet, min_col=2, min_row=2, max_row=17)

    # Add data to chart
    main_chart.add_data(data)
    main_chart.set_categories(labels)
    main_chart.legend = None

    # Chart size
    main_chart.width = 11.8618
    main_chart.height = 9.489

    # Data labels
    main_chart.dataLabels = DataLabelList()
    main_chart.dataLabels.showCatName = True
    main_chart.dataLabels.showPercent = True
    main_chart.dataLabels.showVal = False
    main_chart.dataLabels.showSerName = False

    # Add the chart to the sheet
    chart_sheet.add_chart(main_chart, "C3")

    return main_chart


def create_dichotomy_charts(data_sheet, chart_sheet):
    # Create stacked bar charts for each dichotomy
    create_stacked_dichotomy_chart(data_sheet, chart_sheet, "Energy source - E/I", "D3:D4", "D3:E4", "C24")
    create_stacked_dichotomy_chart(data_sheet, chart_sheet, "Information - S/N", "D7:D8", "D7:E8", "K24")
    create_stacked_dichotomy_chart(data_sheet, chart_sheet, "Decisions - T/F", "D11:D12", "D11:E12", "S24")
    create_stacked_dichotomy_chart(data_sheet, chart_sheet, "Lifestyle - J/P", "D15:D16", "D15:E16", "AA24")


def create_stacked_dichotomy_chart(data_sheet, chart_sheet, title, labels_range, data_range, position):
    # Create a horizontal stacked bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.style = 10
    chart.width = 11.86
    chart.height = 4.24

    # Data and categories
    data = Reference(data_sheet, range_string=f"Data!{data_range}")
    labels = Reference(data_sheet, range_string=f"Data!{labels_range}")
    chart.add_data(data, from_rows=True, titles_from_data=True)
    chart.set_categories(labels)
    chart.legend = Legend()
    chart.legend.position = "t"

    # Stacked, 100%
    chart.grouping = "percentStacked"
    chart.overlap = 100

    # Remove ALL axes
    chart.y_axis.visible = False
    chart.x_axis.visible = False
    chart.y_axis.majorTickMark = "none"
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.delete = True
    chart.x_axis.delete = True
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.ln = LineProperties(noFill=True)
    # Remove gridlines
    chart.x_axis.majorGridlines = None
    chart.x_axis.minorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.y_axis.minorGridlines = None

    # Data labels ON bar only
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = False
    chart.dataLabels.showVal = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showSerName = False

    # Set custom fill colors
    colors = ["4472C4", "C0504D"]  # Blue, Red (hex)
    for idx, ser in enumerate(chart.series):
        ser.graphicalProperties.solidFill = colors[idx]

    # Add chart to sheet
    chart_sheet.add_chart(chart, position)

    return chart


def create_external_internal_charts(data_sheet, chart_sheet):
    # Create Internal Analysis Pie Chart
    internal_pie = PieChart()
    internal_pie.title = "Internal letters Distribution"
    labels = Reference(data_sheet, min_col=7, min_row=3, max_row=6)
    data = Reference(data_sheet, min_col=8, min_row=3, max_row=6)
    internal_pie.add_data(data, titles_from_data=False)
    internal_pie.set_categories(labels)
    internal_pie.width = 11.8618
    internal_pie.height = 9.489

    # Create External Analysis Pie Chart
    external_pie = PieChart()
    external_pie.title = "External letters Distribution"
    labels = Reference(data_sheet, min_col=10, min_row=3, max_row=6)
    data = Reference(data_sheet, min_col=11, min_row=3, max_row=6)
    external_pie.add_data(data, titles_from_data=False)
    external_pie.set_categories(labels)
    external_pie.width = 11.8618
    external_pie.height = 9.489

    # Configure data labels for both charts
    for pie in [internal_pie, external_pie]:
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showCatName = False
        pie.dataLabels.showVal = False
        pie.dataLabels.showPercent = True
        pie.dataLabels.showSerName = False

    # Add the charts to the sheet
    chart_sheet.add_chart(internal_pie, "S3")
    chart_sheet.add_chart(external_pie, "AA3")


def create_dominant_chart(data_sheet, chart_sheet):
    dominant_pie = PieChart()
    dominant_pie.title = "Dominant Function"
    labels = Reference(data_sheet, min_col=7, min_row=11, max_row=18)
    data = Reference(data_sheet, min_col=8, min_row=11, max_row=18)
    dominant_pie.add_data(data, titles_from_data=False)
    dominant_pie.set_categories(labels)
    dominant_pie.width = 11.8618
    dominant_pie.height = 9.489
    # Configure data labels for dominant function chart
    dominant_pie.dataLabels = DataLabelList()
    dominant_pie.dataLabels.showCatName = True
    dominant_pie.dataLabels.showVal = False
    dominant_pie.dataLabels.showPercent = True
    dominant_pie.dataLabels.showSerName = False

    # Add the chart to the sheet
    chart_sheet.add_chart(dominant_pie, "K3")


def create_facet_bar_charts(data_sheet, chart_sheet):
    # Create pie charts for facet dichotomies, organized in rows

    # Row 1: E/I facets
    create_facet_bar_chart(data_sheet, chart_sheet, "D20", "D21:D23", "D21:E23", "C33")  # Initiating-Receiving
    create_facet_bar_chart(data_sheet, chart_sheet, "D26", "D27:D29", "D27:E29", "C40")  # Expressive-Contained
    create_facet_bar_chart(data_sheet, chart_sheet, "D30", "D31:D33", "D31:E33", "C47")  # Gregarious-Intimate
    create_facet_bar_chart(data_sheet, chart_sheet, "D36", "D37:D39", "D37:E39", "C54")  # Active-Reflective
    create_facet_bar_chart(data_sheet, chart_sheet, "D42", "D43:D45", "D43:E45", "C61")  # Enthusiastic-Quiet

    # Row 2: S/N facets
    create_facet_bar_chart(data_sheet, chart_sheet, "D48", "D49:D51", "D49:E51", "K33")  # Concrete-Abstract
    create_facet_bar_chart(data_sheet, chart_sheet, "D54", "D55:D57", "D55:E57", "K40")  # Realistic-Imaginative
    create_facet_bar_chart(data_sheet, chart_sheet, "D60", "D61:D63", "D61:E63", "K47")  # Practical-Conceptual
    create_facet_bar_chart(data_sheet, chart_sheet, "D66", "D67:D69", "D67:E69", "K54")  # Experiential-Theoretical
    create_facet_bar_chart(data_sheet, chart_sheet, "D72", "D73:D75", "D73:E75", "K61")  # Traditional-Original

    # Row 3: T/F facets
    create_facet_bar_chart(data_sheet, chart_sheet, "D78", "D79:D81", "D79:E81", "S33")  # Logical-Empathetic
    create_facet_bar_chart(data_sheet, chart_sheet, "D84", "D85:D87", "D85:E87", "S40")  # Reasonable-Compassionate
    create_facet_bar_chart(data_sheet, chart_sheet, "D90", "D91:D93", "D91:E93", "S47")  # Questioning-Accommodating
    create_facet_bar_chart(data_sheet, chart_sheet, "D96", "D97:D99", "D97:E99", "S54")  # Critical-Accepting
    create_facet_bar_chart(data_sheet, chart_sheet, "D102", "D103:D105", "D103:E105", "S61")  # Tough-Tender

    # Row 4: J/P facets
    create_facet_bar_chart(data_sheet, chart_sheet, "D108", "D109:D111", "D109:E111", "AA33")  # Systematic-Casual
    create_facet_bar_chart(data_sheet, chart_sheet, "D114", "D115:D117", "D115:E117", "AA40")  # Planful-Open-Ended
    create_facet_bar_chart(data_sheet, chart_sheet, "D120", "D121:D123", "D121:E123", "AA47")  # Early Starting-Pressure-Prompted
    create_facet_bar_chart(data_sheet, chart_sheet, "D126", "D127:D129", "D127:E129", "AA54")  # Scheduled-Spontaneous
    create_facet_bar_chart(data_sheet, chart_sheet, "D132", "D133:D135", "D133:E135", "AA61")  # Methodical-Emergent


def create_facet_bar_chart(data_sheet, chart_sheet, title_cell, labels_range, data_range, position):
    """Create a stacked bar chart for a facet dichotomy with improved styling"""
    # Create a horizontal stacked bar chart
    chart = BarChart()
    chart.type = "bar"

    # Get the title from the specified cell
    title = data_sheet[title_cell].value
    chart.title = title
    chart.style = 10

    # Set chart size
    chart.width = 11.8618
    chart.height = 3.175

    # Reference the data and labels
    data = Reference(data_sheet, range_string=f"Data!{data_range}")
    labels = Reference(data_sheet, range_string=f"Data!{labels_range}")
    # Add data and categories
    chart.add_data(data, from_rows=True, titles_from_data=True)
    chart.set_categories(labels)
    chart.legend = Legend()
    chart.legend.position = "t"
    # Make it stacked
    chart.grouping = "percentStacked"
    chart.overlap = 100

    # Remove ALL axes and legend
    chart.y_axis.visible = False
    chart.x_axis.visible = False
    chart.y_axis.majorTickMark = "none"
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.delete = True
    chart.x_axis.delete = True

    # Remove gridlines
    chart.x_axis.majorGridlines = None
    chart.x_axis.minorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.y_axis.minorGridlines = None

    # Data labels ON bar only
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = True    # Show only the category name (label)
    chart.dataLabels.showVal = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.position = "ctr"

    # Remove vertical lines by setting the x-axis line properties to None
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.ln = LineProperties(noFill=True)

    # Add data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = False
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.position = "ctr"
    # Format the percentage display
    chart.dataLabels.numFmt = '0%'
    # Add the chart to the sheet
    chart_sheet.add_chart(chart, position)
    return chart


def create_legend_for_facet_graphs(chart_sheet):
    # Create a dummy bar chart
    chart = BarChart()
    chart.title = "Legend for Facet Graphs"
    chart.style = 10

    # Create dummy data and categories
    data = Reference(chart_sheet, min_col=1, min_row=1, max_row=1)  # Dummy reference
    labels = Reference(chart_sheet, min_col=1, min_row=1, max_row=1)  # Dummy reference
    chart.add_data(data, from_rows=True)
    chart.set_categories(labels)

    # Set legend entries manually
    chart.series = [
        SeriesLabel("In preference (1st facet)"),
        SeriesLabel("In preference (2nd facet)"),
        SeriesLabel("Out of preference (1st facet)"),
        SeriesLabel("Out of preference (2nd facet)"),
        SeriesLabel("MIDZONE")
    ]

    # Remove all axes and gridlines
    chart.y_axis.visible = False
    chart.x_axis.visible = False
    chart.legend.position = 'b'  # Position legend at the bottom
    chart.y_axis.majorTickMark = "none"
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.delete = True
    chart.x_axis.delete = True
    chart.x_axis.majorGridlines = None
    chart.x_axis.minorGridlines = None
    chart.y_axis.majorGridlines = None
    chart.y_axis.minorGridlines = None

    # Add the chart to the sheet
    chart_sheet.add_chart(chart, "O22")


def adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width


def reorder_sheets(workbook):
    """
    Reorders the sheets in the workbook to a logical order:
    1. Main data sheet (MBTI Results)
    2. Charts/Dashboard
    3. Section sheets
    4. Any other sheets
    """
    # Define the preferred order
    preferred_order = ["Dashboard", "MBTI Results", "Charts"]

    # Get current sheet names
    current_sheets = workbook.sheetnames

    # Create a new order list
    new_order = []

    # First add sheets from preferred order if they exist
    for sheet_name in preferred_order:
        if sheet_name in current_sheets:
            new_order.append(sheet_name)
            current_sheets.remove(sheet_name)

    # Then add any section sheets (they typically start with "Section")
    section_sheets = [name for name in current_sheets if name.startswith("Section")]
    section_sheets.sort()  # Sort section sheets numerically if possible
    new_order.extend(section_sheets)
    for name in section_sheets:
        current_sheets.remove(name)

    # Add any remaining sheets
    new_order.extend(current_sheets)

    # Reorder the sheets
    for i, sheet_name in enumerate(new_order):
        sheet = workbook[sheet_name]
        workbook.move_sheet(sheet, i)

    print(f"Sheets reordered to: {new_order}")
    return workbook


def reset_column_widths(chart_sheet):
    """Reset all column widths to default before applying specific widths"""
    # Standard default column width in Excel
    default_width = 8.43  # This is Excel's default column width

    # Reset all columns to default width
    for col_idx in range(1, 100):  # Adjust range as needed based on your dashboard width
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        chart_sheet.column_dimensions[col_letter].width = default_width


if __name__ == "__main__":
    # For testing purposes
    import openpyxl

    workbook_path = r"C:\Users\user\Downloads\group_report_all_pdfs (8).xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    create_distribution_charts(workbook)
    workbook.save(workbook_path)