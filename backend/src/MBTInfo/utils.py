import re
import os
import fitz
from datetime import datetime
import openpyxl
from typing import Dict, Optional, List, Tuple
# local imports
from consts import MBTI_TYPES, FACETS, MIDZONE_FACETS, dominant_functions, All_Facets

lowercase_facets = [facet.lower() for facet in All_Facets]


def parse_mbti_scores(line: str) -> dict:
    # Use regex to find all pairs of MBTI dimension and score
    pairs = re.findall(r'(\w+)\s+\|\s+(\d+)', line)

    # Create a dictionary from the pairs, converting scores to integers
    return {dimension.lower(): int(score) for dimension, score in pairs}


def find_and_parse_mbti_scores(file_path: str) -> dict:
    pattern = r'\b(EXTRAVERSION|INTUITION|THINKING|PERCEIVING)\s+\|\s+\d+'

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if re.search(pattern, line):
                    return parse_mbti_scores(line)
    except IOError as e:
        print(f"Error reading file: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    return {}  # Return empty dictionary if no matching line is found or if there's an error


def get_mbti_type_from_pdf(file_path: str) -> Optional[str]:
    pdf_document = fitz.open(file_path)
    page_num = 1  # Page number to extract text from
    page = pdf_document[page_num]
    text = page.get_text()
    mbti_types = MBTI_TYPES
    for mbti_type in mbti_types:
        if mbti_type in text:
            pdf_document.close()
            return mbti_type
    pdf_document.close()
    return None


def find_type(file_path: str) -> Optional[str]:
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    for mbti_type in MBTI_TYPES:
        if mbti_type in content:
            return mbti_type
    return None


def get_name(file_path: str) -> Optional[str]:
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            # Skip the first 9 lines
            for _ in range(8):
                next(file, None)

            # Read the 10th line
            tenth_line = next(file, '').strip()
            return tenth_line if tenth_line else None
    except Exception as e:
        print(f"Error reading file: {e}")
        return None


def get_date(file_path: str) -> Optional[str]:
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            # Skip the first 10 lines
            for _ in range(9):
                next(file, None)

            # Read the 11th line
            eleventh_line = next(file, '').strip()
            return eleventh_line if eleventh_line else None
    except Exception as e:
        print(f"Error reading file: {e}")
        return None


def get_dominant(file_path: str) -> Optional[str]:
    mbti_type = find_type(file_path)
    if mbti_type and mbti_type in dominant_functions:
        return dominant_functions[mbti_type]
    return None


def get_all_info(file_path: str) -> Dict[str, Optional[str]]:
    info = {
        'name': get_name(file_path),
        'date': get_date(file_path),
        'type': find_type(file_path),
        'dominant': get_dominant(file_path)
    }
    return info


def convert_scores_to_mbti_dict(scores: Dict[str, int]) -> Dict[str, int]:
    """
    Convert a dictionary of MBTI scores to a dictionary with MBTI letters as keys.

    Args:
    scores (Dict[str, int]): A dictionary with MBTI dimensions as keys and scores as values.

    Returns:
    Dict[str, int]: A dictionary with MBTI letters as keys and scores as values.
    """
    mbti_dict = {'E': 0, 'I': 0, 'S': 0, 'N': 0, 'T': 0, 'F': 0, 'J': 0, 'P': 0}

    for dimension, score in scores.items():
        if dimension == 'extraversion':
            mbti_dict['E'] = score
        elif dimension == 'introversion':
            mbti_dict['I'] = score
        elif dimension == 'sensing':
            mbti_dict['S'] = score
        elif dimension == 'intuition':
            mbti_dict['N'] = score
        elif dimension == 'thinking':
            mbti_dict['T'] = score
        elif dimension == 'feeling':
            mbti_dict['F'] = score
        elif dimension == 'judging':
            mbti_dict['J'] = score
        elif dimension == 'perceiving':
            mbti_dict['P'] = score

    return mbti_dict


def get_input_files(directory):
    return [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.pdf')]


def create_output_directory():
    base_dir = r"F:\projects\MBTInfo\output"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(base_dir, timestamp)
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def collect_preferred_qualities(file_path: str) -> List[str]:
    preferred_qualities = []

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        for i, line in enumerate(lines):
            if 'in-preference' in line.lower():
                if i > 0:
                    # Get the previous line and split it into words
                    prev_line = lines[i - 1].strip()
                    words = prev_line.split()
                    if words:
                        # Add the last word from the previous line
                        preferred_qualities.append(words[-1])
        words_to_remove = {'and', 'I', '|5', '|6', 'harmony', 'spontaneity', 'pole.', 'to', 'your'}
        filtered_qualities = [quality for quality in preferred_qualities if quality.lower() not in words_to_remove]
        filtered_qualities.pop(0)

        preferred_qualities = filtered_qualities

        print(f"Found {len(preferred_qualities)} preferred qualities.")
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {str(e)}")

    return preferred_qualities


def collect_midzone_qualities(file_path: str) -> List[str]:
    midzone_qualities = []

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        for i, line in enumerate(lines):
            if 'midzone' in line.lower():
                if i > 0:
                    # Get the previous line and split it into words
                    prev_line = lines[i - 1].strip()
                    words = prev_line.split()
                    if words:
                        # Add the last word from the previous line
                        midzone_qualities.append(words[-1])
        words_to_remove = {'and', 'would', '|5', '|6', 'cause-and-eﬀect,', 'the', 'with', 'facets.', 'distractions.',
                           'preference', 'many'}
        filtered_qualities = [quality for quality in midzone_qualities if quality.lower() not in words_to_remove]
        filtered_qualities.pop(0)

        # Split qualities based on "-" and flatten the list
        split_qualities = [item.strip() for quality in filtered_qualities for item in quality.split('–')]
        # Remove duplicates while preserving order, case-insensitive
        seen = set()
        midzone_qualities = []
        for quality in split_qualities:
            lower_quality = quality.lower()
            if lower_quality not in seen:
                seen.add(lower_quality)
                midzone_qualities.append(quality)

        print(f"Found {len(midzone_qualities)} midzone qualities.")
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {str(e)}")

    return midzone_qualities


def collect_out_qualities(file_path: str) -> List[str]:
    out_qualities = []

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        for i, line in enumerate(lines):
            if 'out-of-preference' in line.lower():
                if i > 0:
                    # Get the previous line and split it into words
                    prev_line = lines[i - 1].strip()
                    words = prev_line.split()
                    if words:
                        # Add the last word from the previous line
                        out_qualities.append(words[-1])
        words_to_remove = {'and', '|5', '|6', 'use.', 'spontaneity', 'preference'}
        filtered_qualities = [quality for quality in out_qualities if quality.lower() not in words_to_remove]
        filtered_qualities.pop(0)

        # Remove duplicates while preserving order
        out_qualities = list(dict.fromkeys(filtered_qualities))

        print(f"Found {len(out_qualities)} out of preference qualities.")
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {str(e)}")

    return out_qualities


def collect_qualities(file_path: str) -> Tuple[List[str], List[str], List[str]]:
    preferred_qualities = collect_preferred_qualities(file_path)
    midzone_qualities = collect_midzone_qualities(file_path)
    out_qualities = collect_out_qualities(file_path)
    return preferred_qualities, midzone_qualities, out_qualities


def check_communication(file_path: str) -> list[str]:
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            start_marker = "YOUR FACET RESULT COMMUNICATION STYLE ENHANCING YOUR STYLE"
            end_marker = "|10"
            communication_facets = []

            start_index = content.find(start_marker)
            end_index = content.find(end_marker, start_index)

            if start_index != -1 and end_index != -1:
                communication_content = content[start_index + len(start_marker):end_index]
                lines = communication_content.split('\n')
                facets_lower = [facet.lower() for facet in FACETS]
                midzone_lower = [facet.lower() for facet in MIDZONE_FACETS]
                for line in lines:
                    words = line.strip().split()
                    if words and words[0].lower() in (facets_lower + midzone_lower):
                        communication_facets.append(words[0].lower())

            filtered_facets = []
            for facet in communication_facets:
                if not any(facet in other_facet and facet != other_facet for other_facet in communication_facets):
                    filtered_facets.append(facet)

            return filtered_facets
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {str(e)}")
        return []


def check_managing_conflict(file_path: str) -> list[str]:
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            start_marker = "YOUR FACET RESULT CONFLICT MANAGEMENT STYLE ENHANCING YOUR STYLE"
            end_marker = "|13"
            decision_facets = []

            start_index = content.find(start_marker)
            end_index = content.find(end_marker, start_index)

            if start_index != -1 and end_index != -1:
                decision_content = content[start_index + len(start_marker):end_index]
                lines = decision_content.split('\n')
                facets_lower = [facet.lower() for facet in FACETS]
                midzone_lower = [facet.lower() for facet in MIDZONE_FACETS]
                for line in lines:
                    words = line.strip().split()
                    if words and words[0].lower() in (facets_lower + midzone_lower):
                        decision_facets.append(words[0].lower())

            filtered_facets = []
            for facet in decision_facets:
                if not any(facet in other_facet and facet != other_facet for other_facet in decision_facets):
                    filtered_facets.append(facet)

            return filtered_facets
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {str(e)}")
        return []


def check_managing_change(file_path: str) -> list[str]:
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            start_marker = "YOUR FACET RESULT CHANGE MANAGEMENT STYLE ENHANCING YOUR STYLE"
            end_marker = "|12"
            change_facets = []

            start_index = content.find(start_marker)
            end_index = content.find(end_marker, start_index)

            if start_index != -1 and end_index != -1:
                change_content = content[start_index + len(start_marker):end_index]
                lines = change_content.split('\n')
                facets_lower = [facet.lower() for facet in FACETS]
                midzone_lower = [facet.lower() for facet in MIDZONE_FACETS]
                for line in lines:
                    words = line.strip().split()
                    if words and words[0].lower() in (facets_lower + midzone_lower):
                        change_facets.append(words[0].lower())

            filtered_facets = []
            for facet in change_facets:
                if not any(facet in other_facet and facet != other_facet for other_facet in change_facets):
                    filtered_facets.append(facet)

            return filtered_facets
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {str(e)}")
        return []


def reorder_sheets(workbook):
    """Reorder sheets to put Dashboard first, followed by main data sheet"""
    # Check if workbook is a string (file path)
    if isinstance(workbook, str):
        try:
            # Load the workbook from the file path
            wb = openpyxl.load_workbook(workbook)
        except Exception as e:
            print(f"Error loading workbook from path '{workbook}': {e}")
            return None
    else:
        # Use the provided workbook object
        wb = workbook

    # Get the current sheet names
    sheet_names = wb.sheetnames

    # Define the desired order - Dashboard first, then Data, then others
    desired_order = []

    # Add Dashboard if it exists
    if 'Dashboard' in sheet_names:
        desired_order.append('Dashboard')

    # Add Data if it exists
    if 'Data' in sheet_names:
        desired_order.append('Data')

    # Add any other sheets in their original order
    for sheet_name in sheet_names:
        if sheet_name not in desired_order:
            desired_order.append(sheet_name)

    # Reorder the sheets
    wb._sheets = [wb[sheet_name] for sheet_name in desired_order]

    # If workbook was a file path, save the changes back to the file
    if isinstance(workbook, str):
        try:
            wb.save(workbook)
        except Exception as e:
            print(f"Error saving workbook to path '{workbook}': {e}")

    return wb


def load_and_reorder_workbook(file_path: str):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Call the reorder_sheets function with the loaded workbook
        reorder_sheets(workbook)

        # Save the workbook if needed
        workbook.save(file_path)
    except Exception as e:
        print(f"Error loading or processing workbook: {e}")


def count_first_words_on_page(file_path: str, word_list: List[str], page_number: int) -> Dict[str, int]:
    """
    Count occurrences of words from a list when they appear as the first word in a line
    on a specific page of a text file. Only counts if the word starts with a capital letter.

    Args:
        file_path (str): Path to the text file
        word_list (List[str]): List of words to search for
        page_number (int): The page number to search in

    Returns:
        Dict[str, int]: Dictionary with words as keys and their occurrence counts as values
    """
    word_counts = {word.lower(): 0 for word in word_list}

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()

            # Find the page content using page markers
            current_page_marker = f"|{page_number}"
            next_page_marker = f"|{page_number + 1}"

            start_index = content.find(current_page_marker)
            if start_index == -1:
                print(f"Page {page_number} not found in {file_path}")
                return word_counts

            # Move past the page marker
            start_index += len(current_page_marker)

            # Find the end of the page
            end_index = content.find(next_page_marker, start_index)
            if end_index == -1:
                # If next page marker not found, read until the end of the file
                end_index = len(content)

            # Extract the page content
            page_content = content[start_index:end_index]

            # Split into lines and check first word of each line
            lines = page_content.split('\n')
            for line in lines:
                line = line.strip()
                if line and line.split():
                    # Get the first word of the line
                    first_word = line.split()[0]
                    
                    # Check if the first word starts with a capital letter
                    if not first_word[0].isupper():
                        continue
                    
                    # Check if the lowercase version of the first word is in our word list
                    first_word_lower = first_word.lower()
                    for word in word_list:
                        if first_word_lower == word.lower():
                            word_counts[word.lower()] += 1
                            break

        return word_counts

    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return word_counts


def count_first_words_across_pages(file_path: str, word_list: List[str], page_numbers: List[int],
                                   only_non_zero: bool = False) -> Dict[str, int]:
    """
    Count occurrences of words from a list when they appear as the first word in a line
    across multiple pages of a text file. Only counts if the word starts with a capital letter.

    Args:
        file_path (str): Path to the text file
        word_list (List[str]): List of words to search for
        page_numbers (List[int]): List of page numbers to search in
        only_non_zero (bool, optional): If True, returns only words with counts > 0. Defaults to False.

    Returns:
        Dict[str, int]: Dictionary with words as keys and their combined occurrence counts as values
    """
    # Initialize counts dictionary
    combined_counts = {word.lower(): 0 for word in word_list}

    # Process each page and combine the counts
    for page_number in page_numbers:
        page_counts = count_first_words_on_page(file_path, word_list, page_number)
        for word, count in page_counts.items():
            combined_counts[word] += count

    # Filter out zero counts if requested
    if only_non_zero:
        combined_counts = {word: count for word, count in combined_counts.items() if count > 0}

    return combined_counts


def extract_sections_between_marker(file_path: str, start_markers: List[str], end_markers: List[str],
                                    page_list: List[int], occurrence_number: int = 0) -> Dict[str, List[str]]:
    """
    For each item in start_markers, find its occurrence as the first word in a line on each page in page_list
    and extract text until the first occurrence of any item from end_markers that appears as the first word on a line.
    Case-insensitive search for markers.
    Removes the word "midzone" and normalizes spaces in the extracted text.

    Args:
        file_path (str): Path to the text file
        start_markers (List[str]): List of starting markers to search for
        end_markers (List[str]): List of ending markers to search for
        page_list (List[int]): List of page numbers to search in
        occurrence_number (int, optional): Which occurrence to extract (0 for all occurrences, 1 for first, 2 for second, etc.)

    Returns:
        Dict[str, List[str]]: Dictionary with start markers as keys and lists of extracted text sections as values
    """
    results = {marker.lower(): [] for marker in start_markers}

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()

            # Process each page in the list
            for page_num in page_list:
                # Find the start page marker
                start_page_marker = f"|{page_num}"
                start_page_index = content.find(start_page_marker)

                if start_page_index == -1:
                    print(f"Page {page_num} not found in {file_path}")
                    continue

                # Find the end page marker (or end of file if not found)
                end_page_marker = f"|{page_num + 1}"
                end_page_index = content.find(end_page_marker, start_page_index)

                if end_page_index == -1:
                    end_page_index = len(content)

                # Extract content only from this page
                page_content = content[start_page_index:end_page_index]

                # Split the page content into lines for easier processing
                page_lines = page_content.split('\n')

                # Process each start marker
                for start_marker in start_markers:
                    start_marker_lower = start_marker.lower()

                    # Find all occurrences where the marker is the first word in a line
                    marker_occurrences = []
                    for i, line in enumerate(page_lines):
                        line = line.strip()
                        if line and line.split() and line.split()[0].lower() == start_marker_lower:
                            marker_occurrences.append(i)

                    # If we're looking for a specific occurrence and it doesn't exist, skip
                    if occurrence_number > 0 and (len(marker_occurrences) < occurrence_number):
                        continue

                    # Determine which occurrences to process
                    occurrences_to_process = []
                    if occurrence_number == 0:  # Process all occurrences
                        occurrences_to_process = marker_occurrences
                    elif occurrence_number <= len(marker_occurrences):  # Process specific occurrence
                        occurrences_to_process = [marker_occurrences[occurrence_number - 1]]

                    # Process each selected occurrence
                    for start_line_idx in occurrences_to_process:
                        # Extract text from the start marker to the end marker
                        extracted_text = []

                        # Get the text after the start marker in the start line (skip the marker itself)
                        start_line = page_lines[start_line_idx].strip()
                        words = start_line.split()
                        if len(words) > 1:  # If there's text after the marker
                            extracted_text.append(' '.join(words[1:]))

                        # Process subsequent lines until we find an end marker at the beginning of a line
                        end_found = False
                        for i in range(start_line_idx + 1, len(page_lines)):
                            line = page_lines[i].strip()
                            if line and line.split():
                                # Check if the line starts with any end marker
                                first_word = line.split()[0].lower()
                                if any(first_word == end_marker.lower() for end_marker in end_markers):
                                    end_found = True
                                    break
                                extracted_text.append(line)

                        if extracted_text:
                            # Join the extracted lines
                            section_text = " ".join(extracted_text)

                            # Remove the word "midzone" (case-insensitive) with proper word boundaries
                            section_text = section_text.replace("midzone", "")
                            # Normalize spaces (replace multiple spaces with a single space)
                            section_text = re.sub(r'\s+', ' ', section_text).strip()
                            section_text = re.sub(r'\bin-preference\b', '', section_text, flags=re.IGNORECASE)


                            # Add the extracted text to the results for this marker
                            results[start_marker_lower].append(section_text)

        # Remove empty lists from results and restore original case for keys
        final_results = {}
        for marker in start_markers:
            marker_lower = marker.lower()
            if marker_lower in results and results[marker_lower]:
                final_results[marker] = results[marker_lower]

        return final_results

    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return {}


def get_facet_explanations(file_path: str, facet: str) -> str:
    facets = [facet]
    lowercase_facets = [facet.lower() for facet in All_Facets]

    result = extract_sections_between_marker(file_path, facets, lowercase_facets, [5, 6, 7, 8], 2)
    print(f"Explanation for '{facet}'")
    result = result[facet][0]
    result = result.replace("in-preference", "")
    result = result.replace("INTERPRETIVE REPORT MYERS-BRIGGS TYPE INDICATOR® | STEP II™ ADAM POMERANTZ ESTJ", "")
    if result:
        return result
    else:
        return f"Error: Could not find '{facet}'"


def get_three_repeating_explanations(file_path: str) -> dict[str, list[str]]:
    page_range = [9, 11, 12]
    non_zero_counts = count_first_words_across_pages(file_path, All_Facets, page_range, only_non_zero=True)
    three_letter_list = []
    two_letter_list = []
    one_letter_list = []
    for item, key in non_zero_counts.items():
        if key == 3:
            three_letter_list.append(item)
        if key == 2:
            two_letter_list.append(item)
        if key == 1:
            one_letter_list.append(item)
            
    # Create a modified version of extract_sections_between_marker for the special case
    def extract_sections_with_special_handling(file_path, start_markers, end_markers, page_list):
        results = {marker.lower(): [] for marker in start_markers}

        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()

                # Process each page in the list
                for page_num in page_list:
                    # Find the start page marker
                    start_page_marker = f"|{page_num}"
                    start_page_index = content.find(start_page_marker)

                    if start_page_index == -1:
                        print(f"Page {page_num} not found in {file_path}")
                        continue

                    # Find the end page marker (or end of file if not found)
                    end_page_marker = f"|{page_num + 1}"
                    end_page_index = content.find(end_page_marker, start_page_index)

                    if end_page_index == -1:
                        end_page_index = len(content)

                    # Extract content only from this page
                    page_content = content[start_page_index:end_page_index]

                    # Split the page content into lines for easier processing
                    page_lines = page_content.split('\n')

                    # Process each start marker
                    for start_marker in start_markers:
                        start_marker_lower = start_marker.lower()

                        # Find all occurrences where the marker is the first word in a line
                        marker_occurrences = []
                        for i, line in enumerate(page_lines):
                            line = line.strip()
                            if line and line.split() and line.split()[0].lower() == start_marker_lower:
                                marker_occurrences.append(i)

                        # Process each occurrence
                        for start_line_idx in marker_occurrences:
                            # Extract text from the start marker to the end marker
                            extracted_text = []

                            # Get the text after the start marker in the start line (skip the marker itself)
                            start_line = page_lines[start_line_idx].strip()
                            words = start_line.split()
                            if len(words) > 1:  # If there's text after the marker
                                extracted_text.append(' '.join(words[1:]))

                            # Process subsequent lines until we find an end marker at the beginning of a line
                            end_found = False
                            for i in range(start_line_idx + 1, len(page_lines)):
                                line = page_lines[i].strip()
                                if line and line.split():
                                    # Check if the line starts with any end marker
                                    first_word = line.split()[0].lower()
                                    
                                    # Special handling for "tough-tender"
                                    if start_marker_lower == "tough-tender" or start_marker_lower == "tough–tender":
                                        # If the first word is "tough" or "tender", don't consider it an end marker
                                        if first_word in ["tough", "tender"]:
                                            extracted_text.append(line)
                                            continue
                                    
                                    # For all other cases, check if it's an end marker
                                    if any(first_word == end_marker.lower() for end_marker in end_markers):
                                        end_found = True
                                        break
                                    extracted_text.append(line)

                            if extracted_text:
                                # Join the extracted lines
                                section_text = " ".join(extracted_text)

                                # Remove the word "midzone" (case-insensitive) with proper word boundaries
                                section_text = section_text.replace("midzone", "")
                                # Normalize spaces (replace multiple spaces with a single space)
                                section_text = re.sub(r'\s+', ' ', section_text).strip()
                                section_text = re.sub(r'\bin-preference\b', '', section_text, flags=re.IGNORECASE)

                                # Add the extracted text to the results for this marker
                                results[start_marker_lower].append(section_text)

            # Remove empty lists from results and restore original case for keys
            final_results = {}
            for marker in start_markers:
                marker_lower = marker.lower()
                if marker_lower in results and results[marker_lower]:
                    final_results[marker] = results[marker_lower]

            return final_results

        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
            return {}
    
    # Use the special handling function instead of the standard one
    three_repeating_explanation = extract_sections_with_special_handling(file_path, three_letter_list, All_Facets, page_range)

    # Additional cleanup to ensure all instances of "midzone" are removed
    # and handle facets with spaces like "Early Starting"
    for marker, texts in three_repeating_explanation.items():
        cleaned_texts = []
        for text in texts:
            # Remove "midzone" with proper word boundaries and normalize spaces
            cleaned_text = re.sub(r'\bmidzone\b', '', text, flags=re.IGNORECASE)
            cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
            
            # Check for other facet names in the text and truncate
            for facet in All_Facets:
                # Skip the current marker to avoid truncating at itself
                if facet.lower() == marker.lower():
                    continue
                
                # Skip if the facet is part of the current marker (for compound markers)
                # For example, if marker is "tough-tender", skip both "tough" and "tender"
                if "-" in marker.lower() or "–" in marker.lower():
                    marker_parts = re.split(r'[-–]', marker.lower())
                    if facet.lower() in marker_parts:
                        continue
                
                # Special handling for "tough-tender"
                if marker.lower() == "tough-tender" or marker.lower() == "tough–tender":
                    # Don't truncate at "tough" or "tender"
                    if facet.lower() in ["tough", "tender"]:
                        continue
                    
                # Look for facet names that might appear in the text
                # Handle both capitalized and all-caps versions
                facet_pattern = re.compile(r'\b' + re.escape(facet) + r'\b', re.IGNORECASE)
                match = facet_pattern.search(cleaned_text)
                if match:
                    # Truncate the text at the facet name
                    cleaned_text = cleaned_text[:match.start()].strip()
            
            # Ensure there's a space after each period
            cleaned_text = re.sub(r'\.(?=[^\s])', '. ', cleaned_text)
            
            cleaned_texts.append(cleaned_text)
        
        three_repeating_explanation[marker] = cleaned_texts

    return three_repeating_explanation


def get_facet_descriptor(filepath: str, facet: str) -> str:
    """
    Extracts the descriptor paragraph for a facet by locating the LAST occurrence
    of its all-caps label ('TOUGH' or 'EXPRESSIVE–CONTAINED') and capturing the paragraph that follows.
    Only searches within pages 5-9 inclusive.
    Removes any "INTERPRETIVE REPORT" line or the last line if present.
    """
    facet = facet.lower()
    print("facet is:", facet)
    
    # Normalize the facet name by replacing hyphens and en dashes with a standard form
    normalized_facet = facet.replace('-', '').replace('–', '').replace(' ', '')
    
    FACET_TITLES = {
        "tough": "TOUGH",
        "expressivecontained": "EXPRESSIVE–CONTAINED",  # Using normalized key
        "expressive": "EXPRESSIVE",
        "contained": "CONTAINED",
        "questioning": "QUESTIONING",
        "toughtender": "TOUGH–TENDER",
        "tender": "TENDER"
    }

    if normalized_facet not in FACET_TITLES:
        return f"[Error] Only 'tough' and 'expressive-contained' are supported. Got '{facet}'."

    target_facet = FACET_TITLES[normalized_facet]

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        return f"[Error] Could not read file: {e}"

    # Extract content only from pages 5-9
    page_content = ""
    for page_num in range(5, 9):  # Pages 5-9 inclusive (note: range is exclusive at the end)
        page_marker = f"|{page_num}"
        next_page_marker = f"|{page_num + 1}"

        start_index = content.find(page_marker)
        if start_index == -1:
            continue  # Page not found, skip to next page

        # Move past the page marker
        start_index += len(page_marker)

        # Find the end of the page
        end_index = content.find(next_page_marker, start_index)
        if end_index == -1:
            # If next page marker not found, read until the end of the file
            end_index = len(content)

        # Extract and append this page's content
        page_content += content[start_index:end_index] + "\n"

    # Split into lines for processing
    lines = page_content.split('\n')

    # Step 1: Find the last index of the facet title in ALL CAPS
    last_index = None
    for i, line in enumerate(lines):
        if line.strip().upper() == target_facet:
            last_index = i

    if last_index is None:
        return f"[Error] Could not find facet '{facet}' in pages 5-9."

    # Step 2: Capture lines following the last occurrence
    descriptor_lines = []
    for line in lines[last_index + 1:]:
        clean = line.strip()

        # Stop if we hit a new all-caps section/facet
        if re.fullmatch(r'[A-Z ()–\-]{4,}', clean):
            break

        # Skip meta-labels and short attribute lines
        clean = clean.replace('in-preference', '')
        clean = clean.replace('midzone', '')
        clean = clean.replace('out-of-preference', '')
        clean = clean.replace("Communicating about disagreements.", '')
        clean = re.sub(r'\.(?=[^\s])', '. ', clean)

        # Accept prose lines
        if clean:
            descriptor_lines.append(clean)

    if not descriptor_lines:
        return "[Error] No descriptor paragraph found after facet."

    # Step 3: Normalize the paragraph
    full_text = ' '.join(descriptor_lines)
    full_text = re.sub(r'\s+', ' ', full_text).strip()

    # Remove specific phrases from the full text
    full_text = full_text.replace("Communicating about disagreements.", "")
    full_text = re.sub(r'\s+', ' ', full_text).strip()  # Normalize spaces again after removal

    # Remove the "INTERPRETIVE REPORT" line if present, or the last line
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', full_text)
    if sentences:
        last_sentence = sentences[-1]
        if "INTERPRETIVE REPORT MYERS-BRIGGS TYPE INDICATOR" in last_sentence:
            # Remove the "INTERPRETIVE REPORT" text and anything after it from the last sentence
            last_sentence = last_sentence.split("INTERPRETIVE REPORT MYERS-BRIGGS TYPE INDICATOR")[0].strip()
            if last_sentence:  # Only keep the sentence if there's content left
                sentences[-1] = last_sentence
            else:
                sentences.pop()  # Remove the last sentence if nothing remains
    if "expressive" in facet.lower() or "contained" in facet.lower():
        # Remove the last sentence for this specific facet
        if sentences and len(sentences) > 1:
            sentences.pop()
    # Special handling for "Questioning" facet - remove the last 3 words


    # Join sentences and ensure proper ending
    result = '\n'.join(s.strip() for s in sentences if s.strip())
    if result and not result.endswith('.'):
        result += '.'
    if facet.lower() == "questioning":
        words = result.split()
        if len(words) > 3:
            result = ' '.join(words[:-3])
    return result




if __name__ == "__main__":
    # test_file_path = r"F:\projects\MBTInfo\output\textfiles\ADAM-POMERANTZ-267149-e4b6edb5-1a5f-ef11-bdfd-6045bd04b01a_text.txt"
    test_file_path = r"F:\projects\MBTInfo\input\Benjamin-Russu-267149-a214ea9d-d272-ef11-bdfd-000d3a58cdb7.pdf"
    print(get_mbti_type_from_pdf(test_file_path))